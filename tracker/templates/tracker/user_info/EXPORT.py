import pymysql
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill

# MySQL connection
connection = pymysql.connect(
    host="localhost",
    user="root",
    password="varshith123",
    database="tasktracker"
)

# Query to fetch data
query = "SELECT * FROM tracker_project"

# Execute query and write to Excel
with connection.cursor() as cursor:
    cursor.execute(query)
    rows = cursor.fetchall()
    columns = [col[0] for col in cursor.description]

    # Create a new workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Tracker Project Data"

    # Write the header row with styling
    for col_num, column_title in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_num, row_data in enumerate(rows, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Define the range for the table
    table_range = f"A1:{openpyxl.utils.get_column_letter(len(columns))}{len(rows) + 1}"

    # Create a table
    table = Table(displayName="ProjectTable", ref=table_range)

    # Apply table style
    style = TableStyleInfo(
        name="WhiteTableStyleLight8",  # Use a predefined table style
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style
    sheet.add_table(table)

    # Adjust column widths for better readability
    for col_num, column_title in enumerate(columns, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(len(column_title), 15)

    # Save the Excel file
    workbook.save('C:/Users/PC-047/Desktop/Book1.xlsx')

print("Data exported successfully to Book1.xlsx")
