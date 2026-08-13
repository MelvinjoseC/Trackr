import openpyxl

def get_session_user(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return None
    return {
        "employee_id": user_id,
        "name": request.session.get("username", "Guest"),
        "designation": request.session.get("designation", "NO DESIGNATION"),
        "authentication": request.session.get("authentication", "No Role")
    }
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
# --- Styled Excel Export for Project Report ---
def export_project_report_excel(request):

    # Create workbook FIRST
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # --- Summary Sheet for All Projects ---
    from openpyxl.chart import BarChart, Reference

    # Fetch all TrackerTasks for the previous month, or for a single project if ?project= is given
    from .models import TrackerTasks
    from datetime import datetime, timedelta
    import re
    today = datetime.today()
    prev_month = today.month - 1 if today.month > 1 else 12
    prev_year = today.year if today.month > 1 else today.year - 1
    project_param = request.GET.get('project')
    if project_param:
        # Only export the selected project
        tasks = TrackerTasks.objects.filter(date1__year=prev_year, date1__month=prev_month, projects=project_param)
    else:
        tasks = TrackerTasks.objects.filter(date1__year=prev_year, date1__month=prev_month)


    # --- Build summary sheet using ALL projects for the month, regardless of filter ---
    all_tasks = TrackerTasks.objects.filter(date1__year=prev_year, date1__month=prev_month)
    summary_project_map = {}
    for t in all_tasks:
        project = t.projects or "(No Project)"
        if project not in summary_project_map:
            summary_project_map[project] = []
        summary_project_map[project].append(t)

    summary_ws = wb.create_sheet(title="Summary", index=0)
    summary_ws['A1'] = 'Project'
    summary_ws['B1'] = 'Total Hours'
    summary_ws['A1'].font = Font(bold=True)
    summary_ws['B1'].font = Font(bold=True)
    summary_project_names = list(summary_project_map.keys())
    total_hours = []
    for i, project in enumerate(summary_project_names, start=2):
        summary_ws[f'A{i}'] = project
        # Sum hours for all tasks in this project
        hours = sum(float(getattr(t, 'time', 0) or 0) for t in summary_project_map[project])
        summary_ws[f'B{i}'] = hours
        total_hours.append(hours)
    # Add bar chart for all projects
    if summary_project_names:
        chart = BarChart()
        chart.title = "Total Hours per Project"
        chart.y_axis.title = 'Total Hours'
        chart.x_axis.title = 'Project'
        data = Reference(summary_ws, min_col=2, min_row=1, max_row=1+len(summary_project_names))
        cats = Reference(summary_ws, min_col=1, min_row=2, max_row=1+len(summary_project_names))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 20
        summary_ws.add_chart(chart, "D2")

    # --- Group by project for the rest of the sheets (filtered if project_param) ---
    project_map = {}
    for t in tasks:
        project = t.projects or "(No Project)"
        if project_param and project != project_param:
            continue
        if project not in project_map:
            project_map[project] = []
        project_map[project].append(t)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    cell_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    align_center = Alignment(horizontal="center", vertical="center")
    alt_fill = PatternFill("solid", fgColor="E9EDF6")

    # Table headers
    headers = [
        "DWG NO", "SCOPE", "STATUS", "START DATE", "END DATE", "REVISION", "HOURS", "PHASE", "DONE BY", "DESCRIPTION OF WORK", "PROJECT PART"
    ]

    from collections import defaultdict
    import calendar
    for project, rows in project_map.items():
        # --- Monthly Overview Sheet ---
        safe_title = re.sub(r'[^\w\- ]', '', project)[:31] or "Sheet1"
        ws = wb.create_sheet(title=safe_title)

        # Monthly Overview Section (with border)
        ws.merge_cells('A1:K1')
        ws['A1'] = 'MONTHLY OVERVIEW'
        ws['A1'].font = Font(bold=True, color='00BFFF', size=16)
        ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28

        overview_labels = [
            ('PROJECT', project),
            ('PO.NO', ''),
            ('MONTH', rows[0].date1.strftime('%B %Y') if rows and rows[0].date1 else ''),
            ('PREPARED BY', ''),
            ('PREPARED ON', 'dd - mm - yyyy'),
        ]
        # Improved Monthly Overview Table Styling
        overview_label_fill = PatternFill("solid", fgColor="F2F2F2")  # Light gray
        overview_value_fill = PatternFill("solid", fgColor="FFFFFF")  # White
        pink_label = 'FFFF00B7'  # Excel ARGB for #FF00B7 (Hot Pink)
        for i, (label, value) in enumerate(overview_labels):
            row = i + 2
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=11)
            # Label cell
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = Font(bold=True, color=pink_label)
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            label_cell.fill = overview_label_fill
            # Value cell
            value_cell = ws.cell(row=row, column=3, value=value)
            value_cell.font = Font(bold=False, color='FF000000')
            value_cell.alignment = Alignment(horizontal="left", vertical="center")
            value_cell.fill = overview_value_fill
            ws.row_dimensions[row].height = 22
        # Draw a single box border around the whole overview section (rows 2-6, cols 1-11)
        for row in range(2, 7):
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                left = 'medium' if col == 1 else None
                right = 'medium' if col == 11 else None
                top = 'medium' if row == 2 else None
                bottom = 'medium' if row == 6 else None
                cell.border = Border(
                    left=Side(style=left or 'thin', color='B7B7B7'),
                    right=Side(style=right or 'thin', color='B7B7B7'),
                    top=Side(style=top or 'thin', color='B7B7B7'),
                    bottom=Side(style=bottom or 'thin', color='B7B7B7')
                )

        ws.row_dimensions[7].height = 10

        ws.merge_cells('A8:E8')
        ws.merge_cells('F8:K8')
        ws['A8'] = 'PROJECT'
        ws['F8'] = 'TOTAL PROJECT HOURS'
        ws['A8'].font = ws['F8'].font = Font(bold=True, color='FF003366')
        ws['A8'].alignment = ws['F8'].alignment = align_center
        for col in range(1, 12):
            cell = ws.cell(row=8, column=col)
            cell.fill = overview_label_fill if col <= 5 else overview_value_fill
            cell.border = Border(
                left=Side(style='thin', color='B7B7B7'),
                right=Side(style='thin', color='B7B7B7'),
                top=Side(style='thin', color='B7B7B7'),
                bottom=Side(style='thin', color='B7B7B7')
            )

        ws.merge_cells('A9:E9')
        ws.merge_cells('F9:K9')
        ws['A9'] = ''
        ws['F9'] = ''
        for col in range(1, 12):
            cell = ws.cell(row=9, column=col)
            cell.fill = overview_value_fill
            cell.border = Border(
                left=Side(style='thin', color='B7B7B7'),
                right=Side(style='thin', color='B7B7B7'),
                top=Side(style='thin', color='B7B7B7'),
                bottom=Side(style='thin', color='B7B7B7')
            )

        ws.merge_cells('A10:E10')
        ws.merge_cells('F10:K10')
        ws['A10'] = 'DESIGNATION'
        ws['F10'] = 'TOTAL HOURS'
        ws['A10'].font = ws['F10'].font = Font(bold=True, color='FF003366')
        ws['A10'].alignment = ws['F10'].alignment = align_center
        for col in range(1, 12):
            cell = ws.cell(row=10, column=col)
            cell.fill = overview_label_fill if col <= 5 else overview_value_fill
            cell.border = Border(
                left=Side(style='thin', color='B7B7B7'),
                right=Side(style='thin', color='B7B7B7'),
                top=Side(style='thin', color='B7B7B7'),
                bottom=Side(style='thin', color='B7B7B7')
            )

        ws.merge_cells('A11:E11')
        ws.merge_cells('F11:K11')
        ws['A11'] = 'PROJECT PART'
        ws['F11'] = 'TOTAL HOURS'
        ws['A11'].font = ws['F11'].font = Font(bold=True, color='FF003366')
        ws['A11'].alignment = ws['F11'].alignment = align_center
        for col in range(1, 12):
            cell = ws.cell(row=11, column=col)
            cell.fill = overview_label_fill if col <= 5 else overview_value_fill
            cell.border = Border(
                left=Side(style='thin', color='B7B7B7'),
                right=Side(style='thin', color='B7B7B7'),
                top=Side(style='thin', color='B7B7B7'),
                bottom=Side(style='thin', color='B7B7B7')
            )

        ws.merge_cells('A12:F12')
        ws.merge_cells('G12:K12')
        ws['A12'] = 'TOTAL HOURS APPROVED'
        ws['G12'] = 'TOTAL HOURS SPENT TO DATE'
        ws['A12'].font = ws['G12'].font = Font(bold=True, color='FF003366')
        ws['A12'].alignment = ws['G12'].alignment = align_center
        for col in range(1, 12):
            cell = ws.cell(row=12, column=col)
            cell.fill = overview_label_fill if col <= 6 else overview_value_fill
            cell.border = Border(
                left=Side(style='thin', color='B7B7B7'),
                right=Side(style='thin', color='B7B7B7'),
                top=Side(style='thin', color='B7B7B7'),
                bottom=Side(style='thin', color='B7B7B7')
            )

        ws.merge_cells('A13:F13')
        ws.merge_cells('G13:K13')
        ws['A13'] = ''
        ws['G13'] = ''
        for col in range(1, 12):
            cell = ws.cell(row=13, column=col)
            cell.fill = overview_value_fill
            cell.border = Border(
                left=Side(style='thin', color='B7B7B7'),
                right=Side(style='thin', color='B7B7B7'),
                top=Side(style='thin', color='B7B7B7'),
                bottom=Side(style='thin', color='B7B7B7')
            )

        ws.merge_cells('A14:K14')
        ws['A14'] = 'SIGNED BY'
        ws['A14'].font = Font(bold=True)
        ws['A14'].alignment = Alignment(horizontal="left")
        for col in range(1, 12):
            ws.cell(row=14, column=col).border = cell_border

        ws.merge_cells('A15:K15')
        ws['A15'] = 'NAME: ' + '_'*30
        ws['A15'].alignment = Alignment(horizontal="left")
        for col in range(1, 12):
            ws.cell(row=15, column=col).border = cell_border

        ws.merge_cells('A16:K16')
        ws['A16'] = 'DATE: ' + '_'*30
        ws['A16'].alignment = Alignment(horizontal="left")
        for col in range(1, 12):
            ws.cell(row=16, column=col).border = cell_border

        ws.merge_cells('A17:K17')
    
        ws['A17'].font = Font(italic=True, color='FF000000')
        ws['A17'].alignment = Alignment(horizontal="left")
        for col in range(1, 12):
            ws.cell(row=17, column=col).border = cell_border

        # --- Data Table Section (Only full month data, no weekly data, no weekly summary at end) ---
        start_row = 19
        
      

        widths = [12, 18, 12, 12, 12, 10, 8, 10, 12, 24, 14]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # --- Week Sheets ---
        # Group rows by week number and create separate sheets (not shown in main/first page)
        week_map = defaultdict(list)
        for t in rows:
            if t.date1:
                week_num = ((t.date1.day - 1) // 7) + 1
                week_map[week_num].append(t)

        from openpyxl.chart import BarChart, Reference
        for week_num, week_rows in week_map.items():
            if not week_rows:
                continue
            # Get week date range
            week_dates = [t.date1 for t in week_rows if t.date1]
            week_start = min(week_dates)
            week_end = max(week_dates)
            week_label = f"WEEK {week_num} ({week_start.strftime('%d-%b')} - {week_end.strftime('%d-%b')})"
            week_sheet_title = f"Week{week_num}_{week_start.strftime('%d%b')}"
            week_sheet_title = week_sheet_title[:31]
            ws_week = wb.create_sheet(title=week_sheet_title)

            # --- Custom Header Table (Rows 1-3) ---
            ws_week.merge_cells('A1:B1')
            ws_week.merge_cells('C1:E1')
            ws_week.merge_cells('F1:K1')
            ws_week['A1'] = 'PROJECT'
            ws_week['C1'] = project
            ws_week['F1'] = ''
            for col in range(1, 12):
                cell = ws_week.cell(row=1, column=col)
                cell.border = Border(left=Side(style='thin', color='000000'),
                                    right=Side(style='thin', color='000000'),
                                    top=Side(style='thin', color='000000'),
                                    bottom=Side(style='thin', color='000000'))
                if col == 1:
                    cell.font = Font(bold=True, color='0000FF')
                cell.alignment = Alignment(horizontal="left" if col in [1,3] else "center", vertical="center")
            ws_week.merge_cells('A2:B2')
            ws_week.merge_cells('C2:E2')
            ws_week.merge_cells('F2:K2')
            ws_week['A2'] = 'WEEK NO'
            ws_week['C2'] = str(week_num)
            ws_week['F2'] = ''
            for col in range(1, 12):
                cell = ws_week.cell(row=2, column=col)
                cell.border = Border(left=Side(style='thin', color='000000'),
                                    right=Side(style='thin', color='000000'),
                                    top=Side(style='thin', color='000000'),
                                    bottom=Side(style='thin', color='000000'))
                if col == 1:
                    cell.font = Font(bold=True, color='0000FF')
                cell.alignment = Alignment(horizontal="left" if col in [1,3] else "center", vertical="center")
            ws_week.merge_cells('A3:K3')
            ws_week['A3'] = ''
            for col in range(1, 12):
                cell = ws_week.cell(row=3, column=col)
                cell.border = Border(left=Side(style='thin', color='000000'),
                                    right=Side(style='thin', color='000000'),
                                    top=Side(style='thin', color='000000'),
                                    bottom=Side(style='thin', color='000000'))

            # --- Data Table Section ---
            table_start_row = 5
            for col, header in enumerate(headers, start=1):
                cell = ws_week.cell(row=table_start_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4472C4")
                cell.alignment = align_center
                cell.border = cell_border

            # Data rows
            for row_idx, t in enumerate(week_rows, start=table_start_row+1):
                ws_week.append([
                    t.d_no or '', t.scope or '', getattr(t, 'project_status', ''),
                    t.start.strftime('%Y-%m-%d') if t.start else '',
                    t.end.strftime('%Y-%m-%d') if t.end else '',
                    t.rev or '', t.time or '', getattr(t, 'phase', ''),
                    t.assigned or '', t.comments or '', t.list or ''
                ])
                for col in range(1, len(headers) + 1):
                    cell = ws_week.cell(row=row_idx, column=col)
                    cell.border = cell_border
                    if (row_idx % 2) == 1:
                        cell.fill = PatternFill("solid", fgColor="E9EDF6")

            for col, width in enumerate(widths, start=1):
                ws_week.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

            # --- Add Bar Chart for Hours per Scope ---
            if week_rows:
                data_start = table_start_row + 1
                data_end = data_start + len(week_rows) - 1
                # Scope is column 2 (B), Hours is column 7 (G)
                chart = BarChart()
                chart.title = "Hours per Scope"
                chart.y_axis.title = 'Hours'
                chart.x_axis.title = 'Scope'
                data = Reference(ws_week, min_col=7, min_row=table_start_row, max_row=data_end)  # Hours
                cats = Reference(ws_week, min_col=2, min_row=data_start, max_row=data_end)  # Scope
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.height = 7
                chart.width = 16
                ws_week.add_chart(chart, f"B{data_end+2}")

            # --- Footer Section: PHASE, DONE BY, DESCRIPTION OF WORK, PROJECT PART ---
            last_row = table_start_row + len(week_rows) + 1
            ws_week.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=11)
            ws_week.cell(row=last_row, column=1).value = f"PHASE: {getattr(week_rows[0], 'phase', '') if week_rows else ''}"
            ws_week.cell(row=last_row, column=1).alignment = Alignment(horizontal="left")
            ws_week.cell(row=last_row, column=1).font = Font(bold=False, color="0000FF")
            ws_week.merge_cells(start_row=last_row+1, start_column=1, end_row=last_row+1, end_column=11)
            ws_week.cell(row=last_row+1, column=1).value = f"DONE BY: {getattr(week_rows[0], 'assigned', '') if week_rows else ''}"
            ws_week.cell(row=last_row+1, column=1).alignment = Alignment(horizontal="left")
            ws_week.cell(row=last_row+1, column=1).font = Font(bold=False, color="0000FF")
            ws_week.merge_cells(start_row=last_row+2, start_column=1, end_row=last_row+2, end_column=11)
            ws_week.cell(row=last_row+2, column=1).value = f"DESCRIPTION OF WORK: {getattr(week_rows[0], 'comments', '') if week_rows else ''}"
            ws_week.cell(row=last_row+2, column=1).alignment = Alignment(horizontal="left")
            ws_week.cell(row=last_row+2, column=1).font = Font(bold=False, color="0000FF")
            ws_week.merge_cells(start_row=last_row+3, start_column=1, end_row=last_row+3, end_column=11)
            ws_week.cell(row=last_row+3, column=1).value = f"PROJECT PART: {getattr(week_rows[0], 'list', '') if week_rows else ''}"
            ws_week.cell(row=last_row+3, column=1).alignment = Alignment(horizontal="left")
            ws_week.cell(row=last_row+3, column=1).font = Font(bold=False, color="0000FF")

    # Response
    filename = f"project_report_{project_param}.xlsx" if project_param else "project_report.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
import calendar
from datetime import datetime
from django.shortcuts import render
from django.db import connection
import base64
from .models import EmployeeDetails
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import ProjectTacker
from django.db import connection, transaction
import matplotlib

matplotlib.use("Agg")  # Set the backend to avoid GUI errors

import matplotlib.pyplot as plt
import numpy as np
import io
import base64
from django.http import JsonResponse


from django.http import JsonResponse
from .models import EmployeeDetails  # Import the EmployeeDetails model
import json

# Define a global variable

def login(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            username = data.get("username")
            password = data.get("password")
        except json.JSONDecodeError:
            return JsonResponse(
                {"success": False, "message": "Invalid request format."}
            )

        # Use Django ORM to check the credentials in the database
        user = EmployeeDetails.objects.filter(name=username).first()

        if user:
            from django.contrib.auth.hashers import check_password, make_password
            is_valid = False
            if check_password(password, user.password):
                is_valid = True
            elif user.password == password:  # Legacy plain text password fallback
                is_valid = True
                user.password = make_password(password)
                user.save(update_fields=['password'])

            if is_valid:
                # Save user info in the session for further authentication
                request.session["user_id"] = user.employee_id
                request.session["username"] = user.name
                request.session["designation"] = user.designation
                request.session["authentication"] = user.authentication
                return JsonResponse({"success": True, "redirect_url": "/task_dashboard/"})
            else:
                return JsonResponse(
                    {"success": False, "message": "Invalid username or password."}
                )
        else:
            return JsonResponse(
                {"success": False, "message": "Invalid username or password."}
            )

    return render(request, "signin.html")
def report_view_page(request):
    from .models import TrackerTasks, EmployeeDetails  # Import here to avoid circular import if any
    import base64

    # Fetch all data from tracker_project table
    all_tasks = TrackerTasks.objects.all()
    tasks_data = list(all_tasks.values())

    # Prepare user info for sidebar from session
    user_id = request.session.get("user_id")
    name = request.session.get("username", "Guest")
    designation = request.session.get("designation", "NO DESIGNATION")
    image_base64 = None

    if user_id:
        try:
            employee = EmployeeDetails.objects.get(employee_id=user_id)
            designation = employee.designation
            if employee.image:
                image_base64 = base64.b64encode(employee.image).decode("utf-8")
        except EmployeeDetails.DoesNotExist:
            designation = "Employee not found"

    # If AJAX/JSON request, return JSON
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        return JsonResponse({"tracker_project_data": tasks_data}, safe=False)

    # Otherwise, render the template as usual, always provide range_6 for template context
    return render(
        request,
        'report_view.html',
        {
            "tracker_project_data": tasks_data,
            "range_6": range(6),
            "name": name,
            "designation": designation,
            "image_base64": image_base64,
            "employee_id": user_id,
        }
    )

def get_admins(request):
    admins = EmployeeDetails.objects.filter(authentication__iexact='admin')  # Case-insensitive check
    admin_list = [{"id": admin.employee_id, "name": admin.name} for admin in admins]
    return JsonResponse({"admins": admin_list})

from django.shortcuts import render
from .models import EmployeeDetails  # Import the EmployeeDetails model
import base64

# Assuming user_data is set earlier

def task_dashboard(request):
    # Default data from session
    user_id = request.session.get("user_id")
    name = request.session.get("username", "Guest")
    designation = request.session.get("designation", "NO DESIGNATION")
    image_base64 = None

    if user_id:
        # Use Django ORM to fetch designation and image from the database
        try:
            employee = EmployeeDetails.objects.get(employee_id=user_id)

            # Update designation
            designation = employee.designation

            # If the image exists, convert it to Base64
            if employee.image:
                image_base64 = base64.b64encode(employee.image).decode("utf-8")

        except EmployeeDetails.DoesNotExist:
            # Handle the case where the employee does not exist
            designation = "Employee not found"
    
    # Pass data to the template
    return render(
        request,
        "tasks_dashboard.html",
        {
            "name": name,
            "designation": designation,
            "image_base64": image_base64,  # Base64-encoded image string
            "employee_id": user_id,  # Pass employee_id to the template
        },
    )



def convert_bytes_safe(data):
    """Convert bytes to strings where possible, handle non-UTF-8 bytes gracefully."""
    if isinstance(data, bytes):
        try:
            # Try decoding as UTF-8
            return data.decode("utf-8")
        except UnicodeDecodeError:
            # If decoding fails, return a placeholder or handle it differently
            return f"[binary data: {len(data)} bytes]"  # Or return None to ignore it

    if isinstance(data, dict):
        return {key: convert_bytes_safe(value) for key, value in data.items()}
    if isinstance(data, list):
        return [convert_bytes_safe(item) for item in data]
    return data

from .models import TrackerTasks
def task_dashboard_api(request):
    # Initialize lists to store tasks and employee details
    task_list = []
    employee_details = []
    project_statuses = []  # Default empty list for statuses

    # Retrieve user_data (assuming it's stored in session)
    user_data
    if user_data:
        # Fetch only statuses where sender_name matches user_data
        project_statuses = list(
            ProjectTacker.objects.filter(sender_name=user_data).values_list(
                "status", flat=True
            )
        )

    try:
        with connection.cursor() as cursor:
            # Fetch all tracker tasks
            cursor.execute("SELECT * FROM tracker_project")
            task_columns = [col[0] for col in cursor.description]
            tasks = cursor.fetchall()
            task_list = [dict(zip(task_columns, task)) for task in tasks]

            # Fetch all employee details
            employees = EmployeeDetails.objects.all()
            employee_columns = [field.name for field in EmployeeDetails._meta.fields]
            employee_details = [dict(zip(employee_columns, [getattr(emp, field) for field in employee_columns])) for emp in employees]
            employee_details = convert_bytes_safe(employee_details)

    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Error fetching data: {str(e)}"}
        )

    # Return a single JSON response with all required data
    return JsonResponse(
        {
            "success": request.user.is_authenticated,
            "message": (
                "Authentication required"
                if not request.user.is_authenticated
                else "Data fetched successfully"
            ),
            "status_list": project_statuses,
            "tasks": task_list,
            "employee_details": employee_details,
        }
    )


def sign_up(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            username = data.get("username")
            password = data.get("password")
        except json.JSONDecodeError:
            return JsonResponse(
                {"success": False, "message": "Invalid request format."}
            )

        if not username or not password:
            return JsonResponse(
                {
                    "success": False,
                    "message": "Both username and password are required.",
                }
            )

        user = EmployeeDetails.objects.filter(name=username).first()
        is_valid = False
        authentication = None

        if user:
            from django.contrib.auth.hashers import check_password, make_password
            if check_password(password, user.password):
                is_valid = True
                authentication = user.authentication
            elif user.password == password:  # Legacy plain text password fallback
                is_valid = True
                authentication = user.authentication
                user.password = make_password(password)
                user.save(update_fields=['password'])

        if is_valid and authentication:
            if authentication.lower() == "admin":
                return render(
                    request, "employee_form.html"
                )  # Render the HTML form for admin users
            else:
                return JsonResponse(
                    {
                        "success": False,
                        "message": "Only admin users are allowed to sign up.",
                    }
                )
        else:
            return JsonResponse(
                {"success": False, "message": "Invalid username or password."}
            )

    return JsonResponse({"success": False, "message": "Invalid request method."})


def save_employee_details(request):
    if request.method == "POST":
        name = request.POST.get("name")
        designation = request.POST.get("designation")
        date_joined = request.POST.get("date_joined")
        email = request.POST.get("email")
        phone_number = request.POST.get("phone_number")
        department = request.POST.get("department")

        status = request.POST.get("status", "Active")
        password = request.POST.get("password")
        image = request.FILES.get("image")  # Get the uploaded file

        from django.contrib.auth.hashers import make_password
        hashed_password = make_password(password) if password else ""

        # Save the data to the EmployeeDetails model
        employee = EmployeeDetails(
            name=name,
            designation=designation,
            date_joined=date_joined,
            email=email,
            phone_number=phone_number,
            department=department,
            status=status,
            password=hashed_password,
            image=image.read() if image else None,  # Convert image to binary
        )
        employee.save()

        return JsonResponse(
            {"success": True, "name": employee.name}
        )  # Respond with success
    else:
        return JsonResponse({"success": False, "message": "Invalid request method."})


def fetch_task_dashboard_data(user_id, selected_date_str):
    """
    Fetch the data required for the task dashboard.

    Args:
        user_id (int): The ID of the logged-in user.
        selected_date_str (str): The selected date as a string.

    Returns:
        dict: A dictionary containing designation, selected_date, and monthly_calendar_data.
    """
    # Fetch the user's designation
    designation = None
    try:
        employee = EmployeeDetails.objects.filter(employee_id=user_id).first()
        designation = employee.designation if employee else None
    except Exception as e:
        print(f"Error fetching designation: {e}")

    # Parse the selected date or use the current date
    try:
        selected_date = (
            datetime.strptime(selected_date_str, "%Y-%m-%d").date()
            if selected_date_str
            else datetime.now().date()
        )
    except ValueError:
        selected_date = datetime.now().date()
        print("Invalid date format provided. Defaulting to today's date.")

    # Fetch task data for the selected date
    monthly_calendar_data = []
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT 
                    id, title, scope, date, time, assigned, category, projects, 
                    list, rev, comments, benchmark, d_no, mail_no, ref_no, created, updated, verification_status, task_status, team
                FROM tracker_project
                WHERE date = %s
            """,
                [selected_date],
            )
            rows = cursor.fetchall()
            monthly_calendar_data = [
                {
                    "id": row[0],
                    "title": row[1],
                    "scope": row[2],
                    "date": row[3],
                    "time": row[4],
                    "assigned": row[5],
                    "category": row[6],
                    "project": row[7],
                    "list": row[8],
                    "rev_no": row[9],
                    "comments": row[10],
                    "benchmark": row[11],
                    "d_no": row[12],
                    "mail_no": row[13],
                    "ref_no": row[14],
                    "created": row[15],
                    "updated": row[16],
                    "verification_status": row[17],
                    "task_status": row[18],
                    "team":row[19],
                }
                for row in rows
            ]
    except Exception as e:
        print(f"Error fetching monthly calendar data: {e}")

    return {
        "designation": designation,
        "selected_date": selected_date,
        "monthly_calendar_data": monthly_calendar_data,
    }


# Helper function to execute SQL queries
def execute_query(query, params=None):
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        if query.strip().lower().startswith("select"):
            # Fetch all rows for SELECT queries
            return cursor.fetchall()
        else:
            # For INSERT, UPDATE, DELETE
            return None

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.utils.dateparse import parse_date
from tracker.models import TrackerTasks
import json

@csrf_exempt
def create_task(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            task_data = data.get("taskData", {})  # Static form data
            excel_tasks = data.get("tasks", [])  # Excel rows

            for task in excel_tasks:
                title = task.get("title", "")                # Column 0
                projects = task.get("projects", "")          # Column 1
                scope = task.get("scope", "")                # Column 2
                category = task.get("category", "")          # Column 3
                task_benchmark = task.get("task_benchmark")  # Column 4
                d_no = task.get("d_no", "")                  # Column 5
                rev = task.get("rev", "")                    # Column 6
                start = task.get("start", "")                # Column 7
                end = task.get("end", "")                    # Column 8
                mail_no = task.get("mail_no", "")            # Column 9
                ref_no = task.get("ref_no", "")              # Column 10

                try:
                    task_benchmark = float(task_benchmark) if task_benchmark else None
                except ValueError:
                    task_benchmark = None

                TrackerTasks.objects.create(
                    title=title,
                    projects=projects,
                    scope=scope,
                    category=category,
                    task_benchmark=task_benchmark,
                    d_no=d_no,
                    rev=rev,
                    mail_no=mail_no,
                    ref_no=ref_no,
                    start=parse_date(start) if start else None,
                    end=parse_date(end) if end else None,
                    team=task_data.get("team", ""),
                    list=task_data.get("list", ""),
                )

            return JsonResponse({"message": "Tasks created successfully!"}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request method"}, status=405)


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.utils.dateparse import parse_date
from tracker.models import TrackerTasks
import json

@csrf_exempt
def edit_task(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            task_id = data.get("id")
            if not task_id:
                return JsonResponse({"error": "Task ID is required"}, status=400)

            task = TrackerTasks.objects.get(id=task_id)

            task.title = data.get("title", "")
            task.projects = data.get("project", "")
            task.scope = data.get("scope", "")
            task.category = data.get("category", "")
            task.start = parse_date(data.get("start_date"))
            task.end = parse_date(data.get("end_date"))
            task.rev = data.get("rev_no", "")
            task.d_no = data.get("d_no", "")
            task.task_benchmark = data.get("task_benchmark", None)

            task.save()
            return JsonResponse({"message": "Task updated successfully!"}, status=200)

        except TrackerTasks.DoesNotExist:
            return JsonResponse({"error": "Task not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request method"}, status=405)


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from tracker.models import TrackerTasks  # Use correct import if your model is elsewhere

@csrf_exempt
def get_task_by_title_project_scope(request):
    if request.method == "GET":
        title = request.GET.get("title")
        project = request.GET.get("project")
        scope = request.GET.get("scope")

        if not (title and project and scope):
            return JsonResponse({"error": "Missing title, project, or scope"}, status=400)

        try:
            task = TrackerTasks.objects.get(title=title, projects=project, scope=scope)

            return JsonResponse({
                "title": task.title,
                "list": task.list,
                "project": task.projects,
                "scope": task.scope,
                "priority": task.priority,
                "assigned_to": task.assigned,
                "checker": task.checker,
                "qc_3_checker": task.qc3_checker,
                "category": task.category,
                "start_date": task.start.isoformat() if task.start else '',
                "end_date": task.end.isoformat() if task.end else '',
                "verification_status": task.verification_status,
                "task_status": task.task_status,
                "rev_no": task.rev,
                "d_no": task.d_no,
                "task_benchmark": task.task_benchmark
            })

        except TrackerTasks.DoesNotExist:
            return JsonResponse({"error": "Task not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request method"}, status=405)

from datetime import datetime, timedelta
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import TrackerTasks

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta
from .models import TrackerTasks

@csrf_exempt
def get_hoursheet_data(request):
    try:
        # Parse start_date and end_date from the query params
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        if not start_date or not end_date:
            today = datetime.today()
            start_of_week = today - timedelta(days=today.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            start_date = start_of_week.date()
            end_date = end_of_week.date()
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        tasks = TrackerTasks.objects.filter(date1__range=[start_date, end_date])

        week_rows = {}
        for task in tasks:
            key = (task.projects, task.scope, task.title, task.category or '')
            if key not in week_rows:
                week_rows[key] = {
                    'projects': task.projects,
                    'scope': task.scope,
                    'title': task.title,
                    'category': task.category or '',
                    'comments': {},
                    'mon': 0, 'tue': 0, 'wed': 0, 'thur': 0, 'fri': 0, 'sat': 0, 'sun': 0
                }

            day_index = task.date1.weekday()
            day_keys = ['mon', 'tue', 'wed', 'thur', 'fri', 'sat', 'sun']
            weekday = day_keys[day_index]
            time_val = float(task.time or 0)
            week_rows[key][weekday] += time_val

            if task.comments:
                week_rows[key]['comments'][weekday] = task.comments

        result_data = []
        for row in week_rows.values():
            total_hours = sum([row[d] for d in ['mon', 'tue', 'wed', 'thur', 'fri', 'sat', 'sun']])
            row['total_hours'] = round(total_hours, 2)
            result_data.append(row)

        dropdowns = list(TrackerTasks.objects.values('projects', 'scope', 'title', 'category').distinct())

        return JsonResponse({
            "draw": int(request.GET.get('draw', 1)),
            "recordsTotal": len(result_data),
            "recordsFiltered": len(result_data),
            "data": result_data,
            "dropdowns": dropdowns
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import TrackerTasks  # or your actual model

@csrf_exempt
def delete_timesheet_row(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            project = data.get('projects')
            scope = data.get('scope')
            title = data.get('title')
            category = data.get('category')
            # You can also filter by user if needed

            TrackerTasks.objects.filter(
                projects=project,
                scope=scope,
                title=title,
                category=category
            ).delete()

            return JsonResponse({"status": "success"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)

from datetime import datetime
from django.http import JsonResponse
from .models import TrackerTasks

def get_filter_data(request):
    try:
        # Get required date range
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        if not start_date_str or not end_date_str:
            return JsonResponse({'error': 'Start and end date are required.'}, status=400)

        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()

        # Optional cascading filters
        selected_project = request.GET.get('project')
        selected_scope = request.GET.get('scope')
        selected_task = request.GET.get('task')

        # Filter base queryset
        queryset = TrackerTasks.objects.filter(date1__range=(start_date, end_date))

        if selected_project:
            queryset = queryset.filter(projects=selected_project)
        if selected_scope:
            queryset = queryset.filter(scope=selected_scope)
        if selected_task:
            queryset = queryset.filter(title=selected_task)

        # Return distinct dropdown values based on current filter level
        projects = queryset.values_list('projects', flat=True).distinct()
        scopes = queryset.values_list('scope', flat=True).distinct()
        tasks = queryset.values_list('title', flat=True).distinct()
        categories = queryset.values_list('category', flat=True).distinct()

        return JsonResponse({
            'projects': list(projects),
            'scopes': list(scopes),
            'tasks': list(tasks),
            'categories': list(categories)
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime
from .models import TrackerTasks
import json


@csrf_exempt
def submit_timesheet(request):
    try:
        user_data = get_session_user(request)
        if not user_data:
            return JsonResponse({"error": "User not logged in."}, status=401)

        username = user_data.get("name")
        data = json.loads(request.body.decode("utf-8"))

        for entry in data:
            # Check if entry has the required fields
            if not entry.get('projects') or not entry.get('scope') or not entry.get('title'):
                return JsonResponse({"error": "Missing required fields."}, status=400)

            # Convert date1 to a valid date object if not already
            try:
                date1 = datetime.strptime(entry['date1'], "%Y-%m-%d").date()
            except ValueError:
                return JsonResponse({"error": f"Invalid date format for date1: {entry['date1']}"}, status=400)

            # Check if there's an existing task with the same project, scope, title, and category
            existing_task = TrackerTasks.objects.filter(
                projects=entry['projects'],
                scope=entry['scope'],
                title=entry['title'],
                category=entry.get('category', '')
            ).first()

            # If the task already exists, update its benchmark, drawing number, and revision number
            if existing_task:
                task_benchmark = entry.get('task_benchmark', existing_task.task_benchmark)
                d_no = entry.get('d_no', existing_task.d_no)
                rev = entry.get('rev', existing_task.rev)
            else:
                task_benchmark = entry.get('task_benchmark', 0)
                d_no = entry.get('d_no', '')
                rev = entry.get('rev', '')

            # Update or create the task entry
            TrackerTasks.objects.update_or_create(
                projects=entry['projects'],
                scope=entry['scope'],
                title=entry['title'],
                date1=date1,
                defaults={
                    'time': float(entry['time']),
                    'category': entry.get('category', ''),
                    'comments': entry.get('comments', ''),
                    'assigned': username,
                    'task_benchmark': task_benchmark,
                    'd_no': d_no,
                    'rev': rev
                }
            )

        return JsonResponse({"message": "Timesheet saved successfully."})
    except Exception as e:
        import traceback
        traceback.print_exc()  # Debugging line
        return JsonResponse({"error": str(e)}, status=500)

from django.http import JsonResponse
from django.db.models import Count, F
from .models import LeaveApplication, Attendance

# Global User Data

def generate_pie_chart(request):
    user_data = get_session_user(request)

    # ✅ Ensure user is logged in
    if not user_data:
        return JsonResponse({"error": "User not logged in."}, status=401)

    username = user_data.get("name")  # Get username from global user data

    try:
        # ✅ Fetch Leave Data using ORM
        leave_data = LeaveApplication.objects.filter(
            username=username, status='Approved'
        ).values(
            'leave_type'
        ).annotate(
            count=Count('id')
        )

        # Initialize leave counts
        full_day_leave = 0
        half_day_leave = 0
        work_from_home = 0

        # Distribute leave counts based on leave_type
        for entry in leave_data:
            if entry['leave_type'] == 'Full Day':
                full_day_leave = entry['count']
            elif entry['leave_type'] == 'Half Day':
                half_day_leave = entry['count']
            elif entry['leave_type'] == 'Work From Home':
                work_from_home = entry['count']

        # ✅ Fetch Attendance Data (Redeemed Leaves) using ORM
        redeemed_days = Attendance.objects.filter(
            username=username, redeemed=1
        ).count()

        # ✅ Calculate Total Working Days (Base 15 + Redeemed Leaves)
        total_working_days = 15 + redeemed_days

        # ✅ Calculate Balance Leave Available
        balance_leave = total_working_days - (full_day_leave + (half_day_leave * 0.5))

        return JsonResponse({
            "balance_leave": balance_leave,
            "full_day_leave": full_day_leave,
            "half_day_leave": half_day_leave,
            "work_from_home": work_from_home,
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)




from django.shortcuts import render, redirect
from django.db import connection
import base64
from .models import ProjectTacker


def project_tracker(request):
    user_data = get_session_user(request)

    # Ensure user is logged in
    if not user_data:
        return redirect("login_page")  # Redirect to login if not logged in

    # Fetch user details from global data
    user_id = user_data.get("employee_id", None)
    name = user_data.get("name", "Guest")
    designation = user_data.get("designation", None)
    authentication = user_data.get("authentication", None)
    image_base64 = None

    # If designation or authentication is not found, fetch from DB
    if user_id and (not designation or not authentication):
        try:
            employee = EmployeeDetails.objects.get(employee_id=user_id)
            designation = employee.designation or "No Designation"
            authentication = employee.authentication or "No Role"
            image_base64 = base64.b64encode(employee.image).decode("utf-8") if employee.image else None
        except EmployeeDetails.DoesNotExist:
            pass

    # Fetch only the records where status is 'Pending'
    project_data = ProjectTacker.objects.filter(status="Pending")

    # Flatten all to_approve JSON data into one list with project info attached
    task_list = []
    for project in project_data:
        to_approve = project.to_aproove
        if to_approve:
            if isinstance(to_approve, dict):
                tasks = [to_approve]
            elif isinstance(to_approve, list):
                tasks = to_approve
            else:
                tasks = []

            for task in tasks:
                task_copy = task.copy()
                task_copy["project_name"] = project.name
                task_copy["sender_name"] = project.sender_name
                task_list.append(task_copy)

    # Determine if user is admin or MD based on the 'authentication' column
    is_admin_or_md = authentication in ['admin', 'MD']

    context = {
        "user_data": user_data,
        "task_list": task_list,
        "name": name,
        "designation": designation,  # <-- Now this is included
        "authentication": authentication,
        "image_base64": image_base64,
        "employee_id": user_id,
        "is_admin_or_md": is_admin_or_md  # <-- Pass it here
    }

    return render(request, "project_tracker.html", context)


import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import ProjectTacker, TrackerTasks

@csrf_exempt
def task_action(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request method."}, status=405)

    try:
        data = json.loads(request.body)
        task_data = data.get("task_data")
        action = data.get("action")

        if not task_data or "name" not in task_data or "d_no" not in task_data:
            return JsonResponse({"success": False, "message": "Task data or identifiers missing."}, status=400)

        project_tracker = ProjectTacker.objects.filter(name=task_data.get("name")).first()
        if not project_tracker:
            return JsonResponse({"success": False, "message": "Project task not found."}, status=404)

        to_approve_data = project_tracker.to_aproove
        if not to_approve_data:
            return JsonResponse({"success": False, "message": "No tasks to approve."}, status=404)

        task = None
        for t in to_approve_data if isinstance(to_approve_data, list) else [to_approve_data]:
            if t.get("d_no") == task_data.get("d_no"):
                task = t
                break

        if not task:
            return JsonResponse({"success": False, "message": "Task not found in project data."}, status=404)

        if action not in ("accept", "reject"):
            return JsonResponse({"success": False, "message": "Invalid action specified."}, status=400)

        project_tracker.status = "Approved" if action == "accept" else "Rejected"
        project_tracker.save()

        if action == "accept":
            tracker_task, created = TrackerTasks.objects.get_or_create(d_no=task.get("d_no"))

            tracker_task.title = task.get("task_title") or task.get("title")
            tracker_task.d_no = task.get("d_no")
            tracker_task.scope = task.get("scope")
            tracker_task.rev = task.get("rev_no")
            tracker_task.checker = task.get("checker")
            tracker_task.projects = task.get("project")
            tracker_task.category = task.get("category")
            tracker_task.end = task.get("end_date")
            tracker_task.priority = task.get("priority")
            tracker_task.start = task.get("start_date")
            tracker_task.assigned = task.get("assigned_to")
            tracker_task.task_status = task.get("task_status") or "Accepted"
            tracker_task.qc3_checker = task.get("qc_3_checker")
            tracker_task.verification_status = task.get("verification_status", True)
            tracker_task.project_status = "Accepted"

            tracker_task.save()

        return JsonResponse({"success": True, "message": f"Task has been {action}ed."})

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON format."}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Error: {str(e)}"}, status=500)


def notifications_view(request):
    return render(request, "notifications.html")


from django.core.exceptions import ObjectDoesNotExist


@csrf_exempt
def check_task_status(request):
    user_data = get_session_user(request)  # Use global variable

    if request.method == "POST":
        try:
            # Fetch all records where sender_name matches user_data['name']
            if not user_data:
                return JsonResponse(
                    {"status": None, "message": "User not authenticated"}, status=401
                )

            sender_name = user_data.get("name")

            projects = ProjectTacker.objects.filter(sender_name__icontains=sender_name)

            if not projects.exists():
                return JsonResponse(
                    {"status": None, "message": "No projects found for sender"},
                    status=404,
                )

            approved_rejected_projects = []

            for project in projects:
                # Convert to_aproove JSONField to a Python dictionary
                to_aprove_data = (
                    json.loads(project.to_aproove)
                    if isinstance(project.to_aproove, str)
                    else project.to_aproove
                )

                if project.status in ["Accepted", "Rejected", "Pending"]:
                    approved_rejected_projects.append(
                        {
                            "status": project.status,
                            "project": to_aprove_data.get("project", "Unknown Project"),
                        }
                    )

            return JsonResponse({"projects": approved_rejected_projects})

        except json.JSONDecodeError:
            return JsonResponse(
                {"status": None, "message": "Invalid JSON format"}, status=400
            )
        except Exception as e:
            return JsonResponse(
                {"status": None, "message": f"Internal server error: {str(e)}"},
                status=500,
            )

    return JsonResponse(
        {"status": None, "message": "Invalid request method"}, status=405
    )


import base64
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.db import connection


def attendance_calendar(request):
    user_data = get_session_user(request)  

    # ✅ Ensure user is logged in
    if not user_data:
        return redirect("login_page")  # Redirect to login if not logged in

    # ✅ Fetch user details from global data
    user_id = user_data.get("employee_id", None)
    name = user_data.get("name", "Guest")
    designation = user_data.get("designation", None)  # Try from global data
    role = user_data.get("role", "").lower()  # Role should be 'admin' or 'user'
    image_base64 = None

    # ✅ If designation is not found in global data, fetch from DB
    if user_id and not designation:
        try:
            employee = EmployeeDetails.objects.get(employee_id=user_id)
            designation = employee.designation or "No Designation"
            image_base64 = base64.b64encode(employee.image).decode("utf-8") if employee.image else None
        except EmployeeDetails.DoesNotExist:
            pass

    # ✅ Check if user is Admin or MD
    is_admin_or_md = False
    employee_auth = EmployeeDetails.objects.filter(name=name).values_list("authentication", flat=True).first()
    if employee_auth:
        auth_value = str(employee_auth).strip().lower()
        is_admin_or_md = (auth_value == "admin" or auth_value == "md")

    # ✅ If request is not GET, return JSON response
    if request.method != "GET":
        return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)

    # ✅ Render template with user details and check for Admin or MD
    return render(request, "calendar.html", {
        "name": name,
        "designation": designation or "No Designation",  # Ensure designation is always present
        "image_base64": image_base64,
        "employee_id": user_id,
        "is_admin_or_md": is_admin_or_md,  # Pass Admin or MD status to template
    })


from django.http import JsonResponse
from django.db import connection

def get_times_by_date(request):
    if request.method == 'GET':
        # Get the date1 parameter from the GET request
        date1 = request.GET.get('date1', None)
        
        if not date1:
            return JsonResponse({'error': 'date1 parameter is required'}, status=400)
        
        try:
            # Fetch rows matching the provided date1
            select_query = """
                SELECT id, title, `list`, projects, scope, category, date1, time, comments, assigned, rev, d_no,
                FROM tracker_project 
                WHERE date1 = %s
                ORDER BY id ASC
            """
            with connection.cursor() as cursor:
                cursor.execute(select_query, [date1])
                rows = cursor.fetchall()
                columns = [col[0] for col in cursor.description]
            
            # Convert the query result to a list of dictionaries
            timesheet_entries = [dict(zip(columns, row)) for row in rows]

            # Return the results as a JSON response
            return JsonResponse({'timesheet_entries': timesheet_entries}, status=200)
        
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)


def get_all_times_by_month(request):
    year = request.GET.get('year')
    month = request.GET.get('month')

    if not year or not month:
        return JsonResponse({'error': 'Year and month parameters are required'}, status=400)

    try:
        # Query to get entries for the given year and month
        query = """
            SELECT id, title, date1, time, projects, scope ,comments, assigned
            FROM tracker_project 
            WHERE YEAR(date1) = %s AND MONTH(date1) = %s
        """
        with connection.cursor() as cursor:
            cursor.execute(query, [year, month])
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]

        # Convert query result to a list of dictionaries
        timesheet_entries = [dict(zip(columns, row)) for row in rows]

        return JsonResponse({'timesheet_entries': timesheet_entries}, status=200)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



from django.http import JsonResponse
from django.db import connection

def get_tasks_by_date(request):
    date1 = request.GET.get('date1')

    if not date1:
        return JsonResponse({'error': 'date1 parameter is required'}, status=400)

    try:
        query = """
            SELECT id, title, projects, scope, date1, time, comments, rev, d_no,
            FROM tracker_project 
            WHERE date1 = %s
        """
        with connection.cursor() as cursor:
            cursor.execute(query, [date1])
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]

        tasks = [dict(zip(columns, row)) for row in rows]
        return JsonResponse({'tasks': tasks}, status=200)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

from django.shortcuts import render
from django.http import JsonResponse
from django.db import connection
from datetime import datetime

def create_project_view(request):
    if request.method == "POST":
        try:
            project_name = request.POST.get("projectName")
            start_date = request.POST.get("startDate")
            end_date = request.POST.get("endDate")
            scope = request.POST.get("scope")
            category = request.POST.get("category")
            benchmark = request.POST.get("Benchmark")

            # Ensure all fields are filled
            if not all([project_name, start_date, end_date, scope, category, benchmark]):
                return JsonResponse({"error": "All fields are required!"}, status=400)

            # Insert into the database
            with connection.cursor() as cursor:
                cursor.execute("""
                    INSERT INTO tracker_project (projects, scope, category, task_benchmark, start, end)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, [project_name, scope, category, benchmark, start_date, end_date])

            return JsonResponse({"message": "Project Created Successfully!"})

        except Exception as e:
            print("Database Error:", e)  # Log the error for debugging
            return JsonResponse({"error": "Something went wrong. Please try again later."}, status=500)

    return render(request, "project_tracker.html")  # Load the form page

from django.shortcuts import render
from django.http import JsonResponse
from django.db import models
from datetime import datetime
from django.core.files.storage import default_storage
from .models import EmployeeDetails, Holiday, LeaveApplication
import base64

# Global user data (Assuming this holds logged-in user info)

def mainleavepage_view(request):
    user_data = get_session_user(request)  

    if not user_data:
        return JsonResponse({"error": "User not logged in."}, status=401)

    # ✅ Fetch user details from global data
    user_id = user_data.get("employee_id", None)
    name = user_data.get("name", "Guest")
    designation = user_data.get("designation", None)  # Try from global data
    image_base64 = None

    # ✅ If designation is not found in global data, fetch from DB
    if user_id and not designation:
        try:
            employee = EmployeeDetails.objects.get(employee_id=user_id)
            designation = employee.designation if employee.designation else "No Designation"
            
            # Check if image exists and is not None
            if employee.image:
                # Check if it's a file path or a bytes object
                if isinstance(employee.image, bytes):
                    image_base64 = base64.b64encode(employee.image).decode("utf-8")
                else:
                    # If it's an ImageField, get the file path and read it
                    with default_storage.open(employee.image.name, 'rb') as image_file:
                        image_base64 = base64.b64encode(image_file.read()).decode("utf-8")
        except EmployeeDetails.DoesNotExist:
            designation = "No Designation"
            image_base64 = None

    today = datetime.today().date()
    current_year = today.year

    # ✅ Fetch holidays
    holidays = Holiday.objects.filter(date__year=current_year).exclude(date__week_day=7).order_by('date')

    # Convert holidays to desired format
    holidays = [{"name": holiday.name, "date": holiday.date.strftime("%Y-%m-%d")} for holiday in holidays]

    # ✅ Fetch leave applications
    leave_applications = LeaveApplication.objects.filter(username=name).values(
        'id', 'start_date', 'end_date', 'reason', 'username', 'approver', 'leave_type', 
        'created_at', 'updated_at', 'status'
    )

    # ✅ Check if user is admin or MD by calling the `check_admin_status` method logic
    auth_result = None
    is_admin = False
    is_md = False

    # Fetch the authentication value from the database for the logged-in user
    try:
        employee = EmployeeDetails.objects.get(name=name)
        if employee.authentication:
            auth_value = employee.authentication.strip().lower()
            is_admin = auth_value == "admin"
            is_md = auth_value == "md"
    except EmployeeDetails.DoesNotExist:
        is_admin = False
        is_md = False

    # Return the context to the template
    return render(request, "mainleavepage.html", {
        "name": name,
        "designation": designation,
        "image_base64": image_base64,
        "employee_id": user_id,
        "is_admin": is_admin,
        "is_md": is_md,
        "holidays": holidays,
        "leave_applications": leave_applications,
    })



from django.shortcuts import render
from django.http import JsonResponse
from datetime import datetime
from .models import LeaveApplication  # Assuming the model is in the same app


def apply_leave_view(request):
    user_data = get_session_user(request)  # Retrieve logged-in user data

    if request.method == "POST":
        try:
            if not user_data:
                return JsonResponse({"error": "User not logged in."}, status=401)

            # ✅ Get the logged-in user's name from user_data
            current_user_name = user_data["name"]
            current_user_id = user_data["employee_id"]

            # ✅ Fetch form data (sent from JavaScript)
            start_date = request.POST.get("from_date", "").strip()
            end_date = request.POST.get("to_date", "").strip()
            leave_type = request.POST.get("leave-type", "").strip()
            reason = request.POST.get("reason", "").strip()
            approver = request.POST.get("approver", "").strip()  # ✅ Now storing the name

            status = "Pending"  # Default status

            # ✅ Debugging: Print received values
            print(f"Received Data - Start: {start_date}, End: {end_date}, Type: {leave_type}, Reason: {reason}, Approver: {approver}")

            # ✅ Validate required fields
            if not all([start_date, end_date, leave_type, reason, approver]):
                return JsonResponse({"error": "All fields are required!"}, status=400)

            # ✅ Create a new leave application using Django ORM
            leave_application = LeaveApplication(
                start_date=start_date,
                end_date=end_date,
                reason=reason,
                username=current_user_name,
                approver=approver,
                leave_type=leave_type,
                status=status
            )

            # ✅ Save the leave application to the database
            leave_application.save()

            return JsonResponse({"message": "Leave request submitted successfully!"})

        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": "Something went wrong. Please try again later."}, status=500)


from django.shortcuts import render
from django.http import JsonResponse
from datetime import datetime
from .models import Holiday  # Import the existing Holiday model

def get_holidays(request):
    """Fetch upcoming holidays only for the current year, excluding Saturdays."""
    try:
        today = datetime.today().date()
        current_year = today.year  # Get current year dynamically

        # Query the Holiday model to get holidays for the current year, excluding Saturdays
        holidays = Holiday.objects.filter(date__year=current_year).exclude(date__week_day=7).order_by('date')

        # Convert holidays to JSON format with status
        holiday_list = []
        for holiday in holidays:
            holiday_date = holiday.date
            status = "past" if holiday_date < today else "upcoming"

            holiday_list.append({
                "name": holiday.name,
                "date": holiday_date.strftime("%Y-%m-%d"),
                "status": status  # Add status for styling in JS
            })

        return JsonResponse({"holidays": holiday_list})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)




from django.http import JsonResponse
from .models import LeaveApplication, Attendance


def leave_application_view(request):
    user_data = get_session_user(request)
    if not user_data:
        return JsonResponse({"error": "User not logged in."}, status=401)

    current_user_name = user_data["name"]

    try:
        # ✅ Fetch Leave Counts by Type (Approved) using Django ORM
        leave_data = LeaveApplication.objects.filter(username=current_user_name, status='Approved').values('leave_type').annotate(count=models.Count('id'))

        # Initialize leave counts
        full_day_leave = 0
        half_day_leave = 0
        work_from_home = 0

        # Distribute leave counts based on leave_type
        for entry in leave_data:
            if entry['leave_type'] == 'Full Day':
                full_day_leave = entry['count']
            elif entry['leave_type'] == 'Half Day':
                half_day_leave = entry['count']
            elif entry['leave_type'] == 'Work From Home':
                work_from_home = entry['count']

        # ✅ Fetch Redeemed Days from Attendance using Django ORM
        redeemed_days = Attendance.objects.filter(username=current_user_name, redeemed=True).count()

        # ✅ Calculate Balance Leave
        total_working_days = 15 + redeemed_days
        balance_leave = total_working_days - (full_day_leave + (half_day_leave * 0.5))

        # ✅ Fetch Leave Applications List using Django ORM
        leave_records = LeaveApplication.objects.filter(username=current_user_name).values(
            'id', 'start_date', 'end_date', 'reason', 'username', 'approver',
            'leave_type', 'created_at', 'updated_at', 'status'
        )

        # ✅ Add "Paid"/"Unpaid" Label Based on balance_leave
        labeled_leaves = []
        for leave in leave_records:
            if balance_leave > 0:
                leave["leave_payment_type"] = "Paid Leave"
                balance_leave -= 1  # Deduct 1 day per leave counted as paid
            else:
                leave["leave_payment_type"] = "Unpaid Leave"
            labeled_leaves.append(leave)

        return JsonResponse({
            "leave_applications": labeled_leaves,
            "balance_leave": balance_leave,
            "full_day_leave": full_day_leave,
            "half_day_leave": half_day_leave,
            "work_from_home": work_from_home,
            "redeemed_days": redeemed_days,
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)



from django.http import JsonResponse
from .models import EmployeeDetails, LeaveApplication  # Import the models


def leave_approvals_view(request):
    user_data = get_session_user(request)
    if not user_data:
        return JsonResponse({"error": "User not logged in."}, status=401)

    # Ensure the logged-in user is an admin
    is_admin = EmployeeDetails.objects.filter(
        name=user_data["name"], authentication__iexact="admin"
    ).exists()

    if not is_admin:
        return JsonResponse({"error": "Forbidden: Only admins can access this."}, status=403)

    # Fetch only pending leave approvals using Django ORM
    pending_leaves = LeaveApplication.objects.filter(status='Pending').values(
        'id', 'start_date', 'end_date', 'reason', 'username', 'approver', 
        'leave_type', 'created_at', 'updated_at', 'status'
    )

    # Convert the queryset to a list of dictionaries
    data = list(pending_leaves)

    return JsonResponse(data, safe=False)



from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import LeaveApplication  # Import the LeaveApplication model
import json

@csrf_exempt  # Temporarily bypass CSRF for debugging (Remove in production)
def update_leave_status(request):
    print("🔍 CSRF Token Received:", request.META.get("HTTP_X_CSRFTOKEN"))  # Debugging

    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method. Only POST is allowed."}, status=405)

    try:
        data = json.loads(request.body)
        leave_id = data.get("id")
        new_status = data.get("status")

        print("🔍 Received Data:", data)  # Debugging

        # Validate input data
        if leave_id is None or new_status is None:
            return JsonResponse({"error": "Missing required fields: 'id' and 'status' are needed."}, status=400)

        # ✅ Update the leave status using the model
        try:
            leave_application = LeaveApplication.objects.get(id=leave_id)
            leave_application.status = new_status
            leave_application.save()  # Save the updated status to the database

            print(f"✅ Leave ID {leave_id} updated to {new_status}")  # Debugging
            return JsonResponse({"message": f"Leave status successfully updated to {new_status}."})

        except LeaveApplication.DoesNotExist:
            return JsonResponse({"error": "Leave application not found."}, status=404)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON data. Ensure the request body is properly formatted."}, status=400)

    except Exception as e:
        print("❌ Unexpected Error:", str(e))  # Debugging
        return JsonResponse({"error": f"Internal Server Error: {str(e)}"}, status=500)



def check_admin_status(request):
    user_data = get_session_user(request)

    if not user_data:
        return JsonResponse({"error": "User not logged in."}, status=401)

    username = user_data.get("name")  # Get logged-in username

    # ✅ Fetch authentication field for the user
    auth_result = EmployeeDetails.objects.filter(name=username).values_list("authentication", flat=True).first()
    
    # ✅ Ensure auth_result is a string and remove spaces
    auth_result = str(auth_result).strip().lower() if auth_result else ""

    print(f"🔍 DEBUG: {username}'s authentication value -> {auth_result}")

    # ✅ Check if the user is Admin or MD
    is_admin = auth_result == "admin"
    is_md = auth_result == "md"

    return JsonResponse({"is_admin": is_admin, "is_md": is_md})


import json
from django.http import JsonResponse
from django.db import connection

def get_task_details(request):
    task_id = request.GET.get("task_id")

    if not task_id:
        return JsonResponse({"error": "Task ID is required"}, status=400)

    try:
        # Fetch task details
        task_query = """
            SELECT id, title, projects, scope, date1, time, comments, list, category, task_status
            FROM tracker_project 
            WHERE id = %s
        """
        with connection.cursor() as cursor:
            cursor.execute(task_query, [task_id])
            row = cursor.fetchone()

        if not row:
            return JsonResponse({"error": "Task not found"}, status=404)

        columns = ["id", "title", "projects", "scope", "date1", "time", "comments", "list", "category", "task_status"]
        task = dict(zip(columns, row))

        # Fetch dropdown values dynamically
        dropdown_data = {}

        # Fetch department lists
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT list FROM tracker_project")
            dropdown_data["list"] = [row[0] for row in cursor.fetchall() if row[0]]

        # Fetch projects based on list
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT list, projects FROM tracker_project")
            dropdown_data["projects"] = [{"list": row[0], "name": row[1]} for row in cursor.fetchall() if row[0] and row[1]]

        # Fetch scopes based on projects
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT projects, scope FROM tracker_project")
            dropdown_data["scope"] = [{"project": row[0], "name": row[1]} for row in cursor.fetchall() if row[0] and row[1]]

        # Fetch tasks based on scopes
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT scope, title FROM tracker_project")
            dropdown_data["titles"] = [{"scope": row[0], "name": row[1]} for row in cursor.fetchall() if row[0] and row[1]]

        # Fetch categories based on tasks
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT title, category FROM tracker_project")
            dropdown_data["category"] = [{"task": row[0], "name": row[1]} for row in cursor.fetchall() if row[0] and row[1]]

        print("Dropdown Data (Backend):", json.dumps(dropdown_data, indent=4))  # Debug print

        return JsonResponse({"task": task, "dropdowns": dropdown_data}, status=200)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)



import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from .models import TrackerTasks

@csrf_exempt
def update_timesheet(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method. Use POST."}, status=405)

    try:
        data = json.loads(request.body.decode("utf-8"))

        task_id = data.get("task_id")
        list_value = data.get("list", "").strip()
        projects = data.get("projects", "").strip()
        scope = data.get("scope", "").strip()
        title = data.get("title", "").strip()
        category = data.get("category", "").strip()
        task_status = data.get("task_status", "").strip()
        date1 = data.get("date1", "").strip()
        comments = data.get("comments", "").strip()

        try:
            time = float(data.get("time", 0))
        except ValueError:
            return JsonResponse({"error": "Invalid value for 'time'. It must be a number."}, status=400)

        if not task_id:
            return JsonResponse({"error": "task_id parameter is required"}, status=400)

        # ✅ Retrieve the task being updated
        try:
            task = TrackerTasks.objects.get(id=task_id)
        except TrackerTasks.DoesNotExist:
            return JsonResponse({"error": "Task not found"}, status=404)

        # ✅ Search for an existing task to fetch team and benchmark
        reference_task = TrackerTasks.objects.filter(
            list=list_value,
            projects=projects,
            scope=scope,
            title=title,
            category=category
        ).exclude(id=task_id).order_by('-id').first()

        # Fallbacks if no matching task found
        team_value = reference_task.team if reference_task else ""
        benchmark_value = reference_task.task_benchmark if reference_task else 0

        # ✅ Update fields
        task.list = list_value
        task.projects = projects
        task.scope = scope
        task.title = title
        task.category = category
        task.task_status = task_status
        task.date1 = date1 if date1 else None
        task.time = time
        task.comments = comments
        task.team = team_value
        task.task_benchmark = benchmark_value

        with transaction.atomic():
            task.save()

        return JsonResponse({"message": "Timesheet updated successfully."}, status=200)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON format"}, status=400)
    except Exception as e:
        print("Error:", str(e))
        return JsonResponse({"error": str(e)}, status=500)


def delete_task(request):
    task_id = request.GET.get('task_id')

    if not task_id:
        return JsonResponse({'error': 'task_id parameter is required'}, status=400)

    try:
        query = "DELETE FROM tracker_project WHERE id = %s"
        with connection.cursor() as cursor:
            cursor.execute(query, [task_id])

        return JsonResponse({'message': 'Task deleted successfully'}, status=200)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

from django.http import JsonResponse
from datetime import datetime
from .models import LeaveApplication  # Import the LeaveApplication model

def delete_leave_application_view(request, leave_id):
    user_data = get_session_user(request)  # Using global variable for user authentication

    if not user_data:
        return JsonResponse({"error": "User not logged in."}, status=401)  # Show as popup

    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method. Use POST instead."}, status=405)

    current_user_name = user_data["name"]  # Get logged-in user's name

    try:
        # Fetch the leave application using Django ORM
        leave = LeaveApplication.objects.get(id=leave_id, username=current_user_name)

        start_date = leave.start_date
        status = leave.status.lower()

        # Check if the start date is in the past
        if start_date < datetime.today().date():
            return JsonResponse({"error": "You cannot delete leave applications with a start date in the past."}, status=403)

        # Delete the leave application
        leave.delete()

        return JsonResponse({"success": "✅ Leave application deleted successfully."})

    except LeaveApplication.DoesNotExist:
        return JsonResponse({"error": "Leave application not found or unauthorized."}, status=404)

    except Exception as e:
        return JsonResponse({"error": f"Internal Server Error: {str(e)}"}, status=500)



from django.http import JsonResponse
from datetime import datetime
from .models import LeaveApplication  # Import the LeaveApplication model

# Global variable to store user data

def edit_leave_application_view(request, leave_id):
    user_data = get_session_user(request)  # Use global variable

    if not user_data:  # Check if user data is available
        return JsonResponse({"error": "User not logged in."}, status=401)

    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=405)

    current_user_name = user_data["name"]  # Get logged-in user's name
    data = request.POST

    # Convert the provided start_date to a datetime object
    try:
        start_date = datetime.strptime(data.get("start_date"), "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"error": "Invalid start date format."}, status=400)

    # Check if the start date is in the past
    if start_date < datetime.today().date():
        return JsonResponse({"error": "You cannot update the leave application to a past date."}, status=403)

    try:
        # Fetch the leave application object
        leave_application = LeaveApplication.objects.get(id=leave_id, username=current_user_name)

        # Update the fields
        leave_application.start_date = start_date
        leave_application.end_date = data.get("end_date")
        leave_application.reason = data.get("reason")
        leave_application.leave_type = data.get("leave_type")
        leave_application.updated_at = datetime.now()  # Update the timestamp

        # Save the updated leave application to the database
        leave_application.save()

        return JsonResponse({"success": "Leave application updated successfully."})

    except LeaveApplication.DoesNotExist:
        return JsonResponse({"error": "Leave application not found or unauthorized."}, status=404)

    except Exception as e:
        return JsonResponse({"error": f"Internal Server Error: {str(e)}"}, status=500)



from django.http import JsonResponse
from datetime import datetime, timedelta
from .models import Attendance, Holiday  # Import the Attendance and Holiday models


def attendance_view(request):
    user_data = get_session_user(request)  # Retrieve logged-in user data

    if request.method == "POST":
        try:
            if not user_data:
                return JsonResponse({"error": "User not logged in."}, status=401)

            # ✅ Get the logged-in user's details
            current_user_name = user_data.get("name")
            current_user_id = user_data.get("employee_id")

            if not current_user_id:
                return JsonResponse({"error": "User ID is missing."}, status=403)

            # ✅ Fetch form data (sent from JavaScript)
            attendance_date = request.POST.get("date", "").strip()
            punch_in = request.POST.get("punch_in", "").strip()
            punch_out = request.POST.get("punch_out", "").strip()
            break_time = request.POST.get("break_time", "").strip()

            # ✅ Debugging: Log Received Data
            print("📢 Received Data:", {
                "date": attendance_date,
                "punch_in": punch_in,
                "punch_out": punch_out,
                "break_time": break_time,
            })

            # ✅ Validate required fields
            if not all([attendance_date, punch_in, punch_out, break_time]):
                return JsonResponse({"error": "All fields are required!"}, status=400)

            # ✅ Convert input values to proper formats
            try:
                attendance_date = datetime.strptime(attendance_date, "%Y-%m-%d").date()
                punch_in = datetime.strptime(punch_in, "%H:%M:%S").time()
                punch_out = datetime.strptime(punch_out, "%H:%M:%S").time()
                break_time = int(break_time)
            except ValueError as e:
                print("📢 Invalid format error:", str(e))  # ✅ Print to Django logs
                return JsonResponse({"error": "Invalid date or time format"}, status=400)

            # ✅ Check if the date is a weekend (Saturday or Sunday) or a holiday
            is_weekend_or_holiday = False
            # Check if it's a weekend (Saturday or Sunday)
            if attendance_date.weekday() == 5 or attendance_date.weekday() == 6:  # 5: Saturday, 6: Sunday
                is_weekend_or_holiday = True
            # Check if it's a holiday
            elif Holiday.objects.filter(date=attendance_date).exists():
                is_weekend_or_holiday = True

            # ✅ Set the `is_compensated` flag based on weekend or holiday check
            is_compensated = 1 if is_weekend_or_holiday else 0

            # ✅ Handle overnight shifts
            dt_punch_in = datetime.combine(attendance_date, punch_in)
            dt_punch_out = datetime.combine(attendance_date, punch_out)

            if dt_punch_out < dt_punch_in:
                dt_punch_out += timedelta(days=1)

            # ✅ Calculate work duration in hours
            work_duration = dt_punch_out - dt_punch_in - timedelta(seconds=break_time)
            work_hours = max(0, work_duration.total_seconds() / 3600.0)  # Convert to hours

            # ✅ Create and save the attendance entry using Django ORM
            attendance = Attendance(
                date=attendance_date,
                punch_in=punch_in,
                punch_out=punch_out,
                break_time=break_time,
                worktime=work_hours,
                user_id=current_user_id,
                is_compensated=is_compensated,  # Set compensation status
                username=current_user_name,
            )
            attendance.save()

            print("✅ Attendance successfully added!")  # ✅ Debugging log

            return JsonResponse({
                "message": "Attendance added successfully!",
                "username": current_user_name,
                "work_hours": work_hours,
                "is_compensated": is_compensated,  # Include compensation status in response
            })

        except Exception as e:
            print("📢 Django Error:", str(e))  # ✅ Log full error in Django console
            return JsonResponse({"error": f"Something went wrong: {str(e)}"}, status=500)

    return JsonResponse({"error": "Invalid request method"}, status=405)


from django.http import JsonResponse
from datetime import datetime, timedelta
from .models import Attendance  # Import the Attendance model


def get_attendance(request):
    user_data = get_session_user(request)

    if not user_data:
        return JsonResponse({"error": "User not logged in."}, status=401)

    user_id = user_data.get("employee_id")

    # Fetch attendance for a specific date
    date = request.GET.get("date", "").strip()
    year = request.GET.get("year", "").strip()
    month = request.GET.get("month", "").strip()

    # Handle requests for monthly attendance
    if year and month:
        try:
            year = int(year)
            month = int(month)
        except ValueError:
            return JsonResponse({}, status=200)  # ✅ Return empty JSON instead of an error

        first_day = datetime(year, month, 1).date()
        last_day = (first_day.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

        # Query using Django ORM
        attendance_records = Attendance.objects.filter(
            user_id=user_id,
            date__range=[first_day, last_day]
        ).order_by('date')

        attendance_data = [
            {
                "date": record.date.strftime("%Y-%m-%d"),
                "punch_in": str(record.punch_in),
                "punch_out": str(record.punch_out),
                "break_time": int(record.break_time),  # Convert break_time to integer
                "worktime": float(record.worktime) if record.worktime is not None else 0.0
            }
            for record in attendance_records
        ]

        return JsonResponse({"attendance": attendance_data}, status=200)

    # Handle single date requests
    if date:
        try:
            date_obj = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            return JsonResponse({}, status=200)  # ✅ Return empty JSON instead of an error

        # Query using Django ORM
        attendance_record = Attendance.objects.filter(
            user_id=user_id,
            date=date_obj
        ).first()

        if attendance_record:
            return JsonResponse({
                "date": attendance_record.date.strftime("%Y-%m-%d"),
                "punch_in": str(attendance_record.punch_in),
                "punch_out": str(attendance_record.punch_out),
                "break_time": int(attendance_record.break_time),
                "worktime": float(attendance_record.worktime) if attendance_record.worktime is not None else 0.0
            }, status=200)

        return JsonResponse({}, status=200)  # ✅ Return empty JSON instead of an error

    return JsonResponse({}, status=200)  # ✅ Return empty JSON instead of an error


from django.http import JsonResponse
from datetime import datetime, timedelta
from .models import Attendance  # Import the Attendance model


def edit_attendance_view(request):
    user_data = get_session_user(request)  # Retrieve logged-in user data

    if request.method == "POST":
        try:
            if not user_data:
                return JsonResponse({"error": "User not logged in."}, status=401)  # ✅ Status 200 prevents redirection

            user_id = user_data.get("employee_id")

            # ✅ Get form data
            attendance_date = request.POST.get("date", "").strip()
            punch_in = request.POST.get("punch_in", "").strip()
            punch_out = request.POST.get("punch_out", "").strip()
            break_time = request.POST.get("break_time", "").strip()
            is_compensated = request.POST.get("is_compensated", "0").strip()  # ✅ Ensure it is always sent

            missing_fields = [field for field, value in {
                "punch_in": punch_in,
                "punch_out": punch_out,
                "break_time": break_time,
                "is_compensated": is_compensated
            }.items() if not value]

            if missing_fields:
                return JsonResponse({"error": f"Missing fields: {', '.join(missing_fields)}"}, status=400)  # ✅ Prevents redirection

            # ✅ Convert input values to correct formats
            try:
                attendance_date = datetime.strptime(attendance_date, "%Y-%m-%d").date()
                punch_in = datetime.strptime(punch_in, "%H:%M:%S").time()
                punch_out = datetime.strptime(punch_out, "%H:%M:%S").time()
                break_time_seconds = int(break_time)  # Ensure break time is stored in seconds
                is_compensated = int(is_compensated)  # Ensure it's a valid integer
            except ValueError:
                return JsonResponse({"error": "Invalid date/time format!"}, status=400)  # ✅ Prevents redirection

            # ✅ Calculate work duration
            dt_punch_in = datetime.combine(attendance_date, punch_in)
            dt_punch_out = datetime.combine(attendance_date, punch_out)

            if dt_punch_out < dt_punch_in:
                dt_punch_out += timedelta(days=1)  # Handle overnight shifts

            work_duration = dt_punch_out - dt_punch_in - timedelta(seconds=break_time_seconds)
            work_hours = max(0, work_duration.total_seconds() / 3600.0)  # Convert to hours

            # ✅ Update attendance record in database using Django ORM
            attendance = Attendance.objects.filter(user_id=user_id, date=attendance_date).first()

            if not attendance:
                return JsonResponse({"error": "Attendance record not found."}, status=404)

            # Update fields
            attendance.punch_in = punch_in
            attendance.punch_out = punch_out
            attendance.break_time = break_time_seconds
            attendance.worktime = work_hours
            attendance.is_compensated = bool(is_compensated)

            # Save the updated record
            attendance.save()

            return JsonResponse({"message": "✅ Attendance updated successfully!", "work_hours": work_hours}, status=200)

        except Exception as e:
            return JsonResponse({"error": f"⚠️ Failed to update attendance. Please try again."}, status=500)  # ✅ Prevents redirection

    return JsonResponse({"error": "Invalid request method"}, status=405)  # ✅ Status 200 prevents error page

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime
from .models import Attendance  # Import the Attendance model


@csrf_exempt
def delete_attendance_view(request):
    user_data = get_session_user(request)

    if request.method == "POST":
        try:
            if not user_data:
                return JsonResponse({"error": "User not logged in."}, status=401)

            user_id = user_data.get("employee_id")
            attendance_date = request.POST.get("date", "").strip()

            if not attendance_date:
                return JsonResponse({"error": "Date is required."}, status=400)

            try:
                attendance_date = datetime.strptime(attendance_date, "%Y-%m-%d").date()
            except ValueError:
                return JsonResponse({"error": "Invalid date format."}, status=400)

            # ✅ Delete the attendance record using Django ORM
            attendance = Attendance.objects.filter(user_id=user_id, date=attendance_date).first()

            if not attendance:
                return JsonResponse({"error": "Attendance record not found."}, status=404)

            # Delete the attendance record
            attendance.delete()

            return JsonResponse({"message": "Attendance deleted successfully."}, status=200)

        except Exception as e:
            return JsonResponse({"error": f"⚠️ Failed to delete attendance. Error: {str(e)}"}, status=500)

    return JsonResponse({"error": "Invalid request method."}, status=405)



from django.http import JsonResponse
from datetime import datetime, timedelta
from .models import Attendance, Holiday, LeaveApplication  # Assuming you have the Attendance, Holiday, and LeaveApplication models


def get_monthly_weekly_attendance(request):
    user_data = get_session_user(request)
    if not user_data:
        return JsonResponse({"error": "User not logged in."}, status=401)

    user_id = user_data.get("employee_id")

    # Get today's date
    today = datetime.today()

    # Get current week (Sunday to Saturday)
    week_start = today - timedelta(days=today.weekday() + 1)  # Get Sunday of this week
    week_end = week_start + timedelta(days=6)  # Get Saturday of this week

    # Get first and last day of the current month
    first_day_of_month = today.replace(day=1)
    last_day_of_month = (first_day_of_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    # Calculate total working days in the month (Monday to Friday)
    total_working_days_month = sum(
        1 for i in range(1, last_day_of_month.day + 1)
        if (first_day_of_month + timedelta(days=i - 1)).weekday() < 5
    )

    # Calculate total working days in the week (Monday to Friday)
    total_working_days_week = sum(
        1 for i in range(7)
        if (week_start + timedelta(days=i)).weekday() < 5  # Monday to Friday only
    )

    # Fetch holidays in the current month (excluding weekends)
    holiday_count_month = Holiday.objects.filter(
        date__range=[first_day_of_month, last_day_of_month]
    ).exclude(date__week_day__in=[6, 7]).count()  # Excluding weekends (Saturday and Sunday)

    # Fetch holidays in the current week (excluding weekends)
    holiday_count_week = Holiday.objects.filter(
        date__range=[week_start, week_end]
    ).exclude(date__week_day__in=[6, 7]).count()  # Excluding weekends (Saturday and Sunday)

    # Fetch leave applications for the current week (excluding weekends)
    leave_applications_week = LeaveApplication.objects.filter(
        username=user_id,
        start_date__gte=week_start,
        end_date__lte=week_end,
        status='Approved'
    )

    # Fetch leave applications for the current month (excluding weekends)
    leave_applications_month = LeaveApplication.objects.filter(
        username=user_id,
        start_date__gte=first_day_of_month,
        end_date__lte=last_day_of_month,
        status='Approved'
    )

    # Calculate leave days taken for the current week (ignoring full/half day)
    leave_taken_week = len(leave_applications_week)

    # Calculate leave days taken for the current month (ignoring full/half day)
    leave_taken_month = len(leave_applications_month)

    # **Expected Monthly Hours Calculation**
    # Total working days in the month minus holidays
    expected_monthly_working_days = total_working_days_month - holiday_count_month
    # Subtract 9 hours for each leave taken
    leave_deduction_month = leave_taken_month * 9
    expected_monthly_hours = expected_monthly_working_days * 9 - leave_deduction_month  # 9 hours per workday

    # **Expected Weekly Hours Calculation**
    # Total working days in the week minus holidays
    expected_weekly_working_days = total_working_days_week - holiday_count_week
    # Subtract 9 hours for each leave taken
    leave_deduction_week = leave_taken_week * 9
    expected_weekly_hours = expected_weekly_working_days * 9 - leave_deduction_week  # 9 hours per workday

    # Fetch total work time for the current month using Django ORM
    total_monthly_hours = Attendance.objects.filter(
        user_id=user_id,
        date__range=[first_day_of_month, last_day_of_month]
    ).aggregate(total_worktime=models.Sum('worktime'))['total_worktime'] or 0.0

    # Fetch total work time for the current week using Django ORM
    total_weekly_hours = Attendance.objects.filter(
        user_id=user_id,
        date__range=[week_start, week_end]
    ).aggregate(total_worktime=models.Sum('worktime'))['total_worktime'] or 0.0

    return JsonResponse({
        "total_monthly_hours": total_monthly_hours,
        "total_weekly_hours": total_weekly_hours,
        "expected_monthly_hours": expected_monthly_hours,
        "expected_weekly_hours": max(expected_weekly_hours, 0)  # Prevents negative values
    })


from django.http import JsonResponse
from .models import Attendance  # Import the Attendance model


def get_compensated_worktime(request):
    """Fetch compensated worktime records for the logged-in user."""
    user_data = get_session_user(request)

    if not user_data:
        return JsonResponse({"error": "User not logged in."}, status=401)

    username = user_data.get("name")  # Get username from global user data

    try:
        # Fetch compensated worktime records using Django ORM
        compensated_worktime_records = Attendance.objects.filter(
            is_compensated=True,
            username=username
        ).order_by('-date')  # Order by date descending

        # Prepare the data to be returned as JSON
        compensated_worktime_data = [
            {
                "id": record.id,
                "date": record.date.strftime("%Y-%m-%d"),
                "punch_in": str(record.punch_in),
                "punch_out": str(record.punch_out),
                "break_time": int(record.break_time),
                "worktime": float(record.worktime) if record.worktime is not None else 0.0,
                "user_id": record.user_id,
                "is_compensated": record.is_compensated,
                "redeemed": record.redeemed,
                "username": record.username
            }
            for record in compensated_worktime_records
        ]

        return JsonResponse({"compensated_worktime": compensated_worktime_data}, status=200)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


from django.http import JsonResponse
from .models import Attendance  # Import the Attendance model
import json


def request_comp_leave(request):
    """User submits a request for compensatory leave approval."""
    user_data = get_session_user(request)

    if not user_data:
        return JsonResponse({"error": "User not logged in."}, status=401)

    if request.method != "POST":
        return JsonResponse({"error": "Method Not Allowed"}, status=405)

    try:
        data = json.loads(request.body)
        worktime_id = data.get("id")

        if not worktime_id:
            return JsonResponse({"error": "Missing worktime ID"}, status=400)

        # Fetch the attendance record with the given worktime_id using Django ORM
        attendance = Attendance.objects.filter(id=worktime_id).first()

        if not attendance:
            return JsonResponse({"error": "Attendance record not found."}, status=404)

        # Mark the compensatory leave request as pending approval (set is_compensated to 2)
        # Using is_compensated as an integer field (0 = not compensated, 2 = pending approval)
        attendance.is_compensated = 2
        attendance.save()  # Save the updated record

        return JsonResponse({"message": "Request submitted for approval."}, status=200)

    except Exception as e:
        return JsonResponse({"error": f"Failed to update attendance. Error: {str(e)}"}, status=500)


from django.http import JsonResponse
from .models import Attendance, EmployeeDetails  # Import the models
import json


def get_pending_comp_leave_requests(request):
    """MD fetches pending compensatory leave requests."""
    user_data = get_session_user(request)

    if not user_data:
        return JsonResponse({"error": "User not logged in."}, status=401)

    username = user_data.get("name")
    is_md = False

    # Check if the user is MD using Django ORM
    try:
        employee = EmployeeDetails.objects.filter(name=username).first()
        if employee and employee.authentication == "MD":
            is_md = True
    except EmployeeDetails.DoesNotExist:
        pass

    if not is_md:
        return JsonResponse({"error": "Forbidden: Only MD can view this."}, status=403)

    try:
        # Fetch pending compensatory leave requests using Django ORM
        pending_requests = Attendance.objects.filter(is_compensated=2).values(
            'id', 'date', 'username', 'worktime'
        )

        # Return the data as a list of dictionaries
        pending_requests_data = list(pending_requests)

        return JsonResponse({"pending_requests": pending_requests_data}, status=200)

    except Exception as e:
        return JsonResponse({"error": f"An error occurred: {str(e)}"}, status=500)

from django.http import JsonResponse
from django.db import transaction
from .models import Attendance, EmployeeDetails  # Import the models
import json


@csrf_exempt
def update_comp_leave_status(request):
    """MD approves or rejects a compensatory leave request."""
    user_data = get_session_user(request)

    if not user_data:
        return JsonResponse({"error": "User not logged in."}, status=401)

    username = user_data.get("name")
    is_md = False

    try:
        # Check if the logged-in user is an MD using Django ORM
        employee = EmployeeDetails.objects.filter(name=username).first()
        if employee and employee.authentication == "MD":
            is_md = True
    except EmployeeDetails.DoesNotExist:
        pass

    if not is_md:
        return JsonResponse({"error": "Forbidden: Only MD can approve/reject."}, status=403)

    if request.method != "POST":
        return JsonResponse({"error": "Method Not Allowed"}, status=405)

    try:
        data = json.loads(request.body)
        record_id = data.get("id")
        action = data.get("action")  # "approve" or "reject"

        if not record_id or action not in ["approve", "reject"]:
            return JsonResponse({"error": "Invalid request data."}, status=400)

        with transaction.atomic():
            # Fetch the attendance record using Django ORM
            attendance = Attendance.objects.filter(id=record_id).first()

            if not attendance:
                return JsonResponse({"error": "Attendance record not found."}, status=404)

            # Update the attendance record based on the action (approve/reject)
            if action == "approve":
                attendance.is_compensated = 0  # Set is_compensated to 0 for approved
                attendance.redeemed = 1  # Set redeemed to 1 for approved
            elif action == "reject":
                attendance.is_compensated = 1  # Set is_compensated to 1 for rejected
                attendance.redeemed = 0 # Set redeemed to 0 for approved
            # Save the updated record
            attendance.save()

        return JsonResponse({"message": f"Comp Leave {action}d successfully."}, status=200)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)




from django.shortcuts import render
from datetime import datetime
from django.db.models import Sum
from .models import EmployeeDetails, Attendance  # Import models

def monthly_attendance_view(request):
    # Get the current month and year
    current_month = datetime.now().month
    current_year = datetime.now().year

    # Fetch the attendance data for the current month and year using Django ORM
    attendance_data = (
        EmployeeDetails.objects
        .filter(
            attendance__date__month=current_month,
            attendance__date__year=current_year
        )
        .annotate(
            total_worktime=Sum('attendance__worktime')  # Calculate the total worktime per employee
        )
        .values('name', 'total_worktime', 'attendance__date', 'attendance__worktime')
        .order_by('name', 'attendance__date')
    )

    # Format the data to match the previous structure
    attendance_data_list = [
        {
            'name': data['name'],
            'total_worktime': data['total_worktime'],
            'date': data['attendance__date'],
            'daily_worktime': data['attendance__worktime']
        }
        for data in attendance_data
    ]

    return render(request, "project_tracker.html", {'attendance_data': attendance_data_list})

from django.http import JsonResponse
from .models import EmployeeDetails  # Import the EmployeeDetails model

def get_employee_names(request):
    # Fetch employee names and ids from EmployeeDetails using Django ORM
    employees = EmployeeDetails.objects.all().order_by('name').values('employee_id', 'name')

    # Prepare the data to be returned in JSON format
    employee_data = [{'id': employee['employee_id'], 'name': employee['name']} for employee in employees]

    return JsonResponse({'employees': employee_data})


from django.http import JsonResponse
from django.db.models import Sum
from .models import Attendance  # Import the Attendance model

def get_user_worktime(request):
    try:
        employee_id = request.GET.get('employee_id')
        if not employee_id:
            return JsonResponse({'error': 'Missing employee_id'}, status=400)

        # Fetch the attendance records for the specific employee using Django ORM
        worktime_data = (
            Attendance.objects
            .filter(user_id=employee_id)  # Filter by the user_id
            .values('date', 'worktime')  # Select the necessary fields: date and worktime
            .annotate(total_worktime=Sum('worktime'))  # Calculate total worktime for the employee
            .order_by('date')  # Order by date
        )

        # Format the data to match the desired output structure
        worktime_data_list = [
            {
                'date': entry['date'],
                'daily_worktime': entry['worktime'],
                'total_worktime': entry['total_worktime']
            }
            for entry in worktime_data
        ]

        return JsonResponse({'worktime': worktime_data_list})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
from django.http import JsonResponse
from .models import Attendance
from django.views.decorators.http import require_GET

@require_GET
def get_attendance_details(request):
    employee_id = request.GET.get('employee_id')
    date = request.GET.get('date')  # Expected format: YYYY-MM-DD

    if not employee_id or not date:
        return JsonResponse({'error': 'Missing employee_id or date'}, status=400)

    try:
        attendance = Attendance.objects.filter(
            user_id=employee_id,
            date=date
        ).values('punch_in', 'punch_out', 'break_time', 'worktime').first()

        if not attendance:
            # ✅ Return status=200 with an error message to avoid 404
            return JsonResponse({'error': 'No data for selected date'}, status=200)

        return JsonResponse({'attendance': attendance}, status=200)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


from django.http import JsonResponse
from .models import TrackerTasks  # Import the TrackerTasks model

def monthly_project_analysis(request):
    project_name = request.GET.get('project_name', None)

    # Start the queryset to fetch data from TrackerTasks model
    project_query = TrackerTasks.objects.all()

    # Apply filtering if project_name is provided
    if project_name:
        project_query = project_query.filter(projects=project_name)

    # Fetch the relevant fields including task_benchmark
    project_data = project_query.values('projects', 'category', 'date1', 'time', 'task_benchmark', 'title').order_by('projects', 'date1')

    # Convert the queryset to a list of dictionaries
    project_data_list = [
        {
            'projects': item['projects'],
            'category': item['category'],
            'date1': item['date1'],
            'time': item['time'],
            'task_benchmark': item['task_benchmark'],
            'title':item['title'],
        }
        for item in project_data
    ]

    return JsonResponse({'projects': project_data_list})



from django.http import JsonResponse
from .models import TrackerTasks  # Import the TrackerTasks model

def get_project_categories(request):
    project_name = request.GET.get('project_name', None)

    if not project_name:
        return JsonResponse({'error': 'No project name provided'}, status=400)

    try:
        # Use Django ORM to get the distinct categories for the given project name
        categories = (
            TrackerTasks.objects
            .filter(projects=project_name)  # Filter by project name
            .values('category')  # Select the category field
            .distinct()  # Get distinct categories
        )

        # Extract the category names from the queryset
        category_list = [category['category'] for category in categories]

        # Return the categories in a JsonResponse
        return JsonResponse({'categories': category_list})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



from django.http import JsonResponse
from .models import TrackerTasks
from datetime import datetime, timedelta

def get_week_date_range(week_offset):
    """
    Helper function to get the date range for the given week offset.
    week_offset: 0 for CW (current week), -1 for W6 (previous week), etc.
    """
    today = datetime.today()
    start_of_week = today - timedelta(days=today.weekday())  # Monday of the current week

    # Calculate the start and end of the target week
    week_start = start_of_week + timedelta(weeks=week_offset)
    week_end = week_start + timedelta(days=6)

    return week_start.date(), week_end.date()
from django.http import JsonResponse
from .models import TrackerTasks
from datetime import datetime, timedelta

def get_week_date_range(week_offset):
    """
    Helper function to get the date range for the given week offset.
    week_offset: 0 for CW (current week), -1 for W6 (previous week), etc.
    """
    today = datetime.today()
    start_of_week = today - timedelta(days=today.weekday())  # Monday of the current week

    # Calculate the start and end of the target week
    week_start = start_of_week + timedelta(weeks=week_offset)
    week_end = week_start + timedelta(days=6)

    return week_start.date(), week_end.date()

def get_project_datas(request):
    department = request.GET.get('department', None)  # Use 'list' here for department
    project_name = request.GET.get('project_name', None)
    category = request.GET.get('category', None)
    week_offset = int(request.GET.get('week_offset', 0))  # Week offset: 0 for CW, -1 for W6, etc.

    # Get the date range for the selected week
    week_start, week_end = get_week_date_range(week_offset)

    # Start the queryset to fetch data from TrackerTasks model
    project_query = TrackerTasks.objects.all()

    # Apply filtering if department is provided
    if department:
        project_query = project_query.filter(list=department)  # Use 'list' field for department
    
    # Apply filtering if project_name is provided
    if project_name:
        project_query = project_query.filter(projects=project_name)

    # Apply filtering if category is provided
    if category:
        project_query = project_query.filter(category=category)
    
    # Apply filtering for the week based on the date range
    project_query = project_query.filter(date1__range=[week_start, week_end])

    # Fetch the relevant fields and annotate results
    project_data = project_query.values('projects', 'category', 'date1', 'time', 'list').order_by('projects', 'date1')

    # Convert the queryset to a list of dictionaries
    project_data_list = [
        {
            'projects': item['projects'],
            'category': item['category'],
            'date1': item['date1'],
            'time': item['time'],
            'department': item['list']
        }
        for item in project_data
    ]
    
    # Get available departments for the department dropdown
    departments = TrackerTasks.objects.values('list').distinct()
    department_list = [department['list'] for department in departments]

    # Get available projects for the project dropdown based on the selected department
    if department:
        projects = TrackerTasks.objects.filter(list=department).values('projects').distinct()
        project_list = [project['projects'] for project in projects]
    else:
        project_list = []

    # Get available categories for the category dropdown based on the selected project
    if project_name:
        categories = TrackerTasks.objects.filter(projects=project_name).values('category').distinct()
        category_list = [category['category'] for category in categories]
    else:
        category_list = []

    # Available weeks (CW, W6, W5, etc.)
    week_list = [
        {'label': 'CW', 'value': 0},
        {'label': 'W6', 'value': -1},
        {'label': 'W5', 'value': -2},
        {'label': 'W4', 'value': -3},
        {'label': 'W3', 'value': -4},
        {'label': 'W2', 'value': -5},
        {'label': 'W1', 'value': -6},
    ]

    # Return the response with department, project, category, week data, and the filtered project data
    return JsonResponse({
        'departments': department_list,
        'projects': project_list,
        'categories': category_list,
        'weeks': week_list,
        'project_data': project_data_list
    })


from django.http import JsonResponse
from .models import TrackerTasks
from datetime import datetime

def get_task_details_for_sidebar(request):
    selected_date = request.GET.get('date')

    if not selected_date:
        return JsonResponse({'error': 'date is required'}, status=400)

    try:
        selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()

        # Query tasks by date
        tasks = TrackerTasks.objects.filter(date1=selected_date)

        if not tasks.exists():  # ✅ Use exists() instead of evaluating queryset
            return JsonResponse({'tasks': []})  # ✅ Return empty list, not error

        # Prepare task data
        task_data = [{
            'title': task.title,
            'projects': task.projects,
            'scope': task.scope,
            'category': task.category,
            'time': task.time,
            'comments': task.comments,
            'task_benchmark': task.task_benchmark,
            'assigned': task.assigned
        } for task in tasks]

        return JsonResponse({'tasks': task_data})

    except ValueError:
        return JsonResponse({'error': 'Invalid date format, expected YYYY-MM-DD'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'An error occurred: {str(e)}'}, status=500)


from django.shortcuts import render
from .models import TrackerTasks, EmployeeDetails  # Your model import
import base64
from django.db import connection

# Assuming user_data is a global variable

def team_dashboard(request):
    user_data = get_session_user(request)  # Access the global variable for user data

    # Default data if user_data is not set
    user_id = user_data.get("employee_id", None) if user_data else None
    name = user_data.get("name", "Guest") if user_data else "Guest"
    designation = user_data.get("designation", None) if user_data else None
    authentication = user_data.get("authentication", None) if user_data else None
    image_base64 = None  # Initialize empty image

    # If designation or authentication is not found, fetch from DB
    if user_id and (not designation or not authentication):
        try:
            employee = EmployeeDetails.objects.get(employee_id=user_id)
            designation = employee.designation or "No Designation"
            authentication = employee.authentication or "No Role"
            image_base64 = base64.b64encode(employee.image).decode("utf-8") if employee.image else None
        except EmployeeDetails.DoesNotExist:
            pass

    # Determine if user is admin or MD based on the 'authentication' column
    is_admin_or_md = authentication in ['admin', 'MD']

    # Fetch TrackerTasks data
    tracker_data = []

    if user_id:
        try:
            # Get all tasks related to the user's team or their assignment
            tasks = TrackerTasks.objects.filter(assigned=name)  # Assuming 'assigned' stores employee name
            for task in tasks:
                tracker_data.append({
                    "team": task.team,
                    "task_benchmark": task.task_benchmark,  # APPROVED HOURS
                    "time": task.time,  # Total Worktime
                    "projects": task.projects,
                    "project_status": task.project_status,
                    "title": task.title,
                    "task_status": task.task_status,
                    "priority": task.priority,
                    "start": task.start,
                    "end": task.end,
                })
        
        except TrackerTasks.DoesNotExist:
            tracker_data = []

    # Now pass the data to the template
    return render(
        request,
        "team_dashboard.html",  # Your team dashboard template
        {
            "name": name,
            "designation": designation,
            "authentication": authentication,
            "image_base64": image_base64,  # Send base64 encoded image
            "employee_id": user_id,  # Pass employee_id to template
            "tracker_data": tracker_data,  # Pass tracker task data to template
            "is_admin_or_md": is_admin_or_md  # <-- Pass it here as well
        },
    )

from django.http import JsonResponse
from .models import TrackerTasks
from django.db.models import Sum

def get_projects(request):
    # Get distinct project names
    distinct_projects = TrackerTasks.objects.values_list('projects', flat=True).distinct()

    # Prepare response containers
    project_names = []
    approved_hours = []
    total_worktime = []

    for project in distinct_projects:
        project_names.append(project)

        tasks = TrackerTasks.objects.filter(projects=project).values(
            'scope', 'category', 'title', 'rev', 'd_no', 'task_benchmark', 'time'
        )

        # Approved hours should be based on unique task definitions
        unique_keys = set()
        approved_sum = 0
        worktime_sum = 0

        for task in tasks:
            key = (task['scope'], task['category'], task['title'], task['rev'], task['d_no'])
            if key not in unique_keys:
                unique_keys.add(key)
                approved_sum += float(task['task_benchmark'] or 0)

            worktime_sum += float(task['time'] or 0)

        approved_hours.append(approved_sum)
        total_worktime.append(worktime_sum)

    # Final combined response
    return JsonResponse({
        "projects": project_names,
        "approvedHours": approved_hours,
        "totalWorktime": total_worktime
    })



# ✅ New view: Get project-level data (approved + total worktime)
def get_project_data(request, project):
    # Filter only for the selected project
    tasks = TrackerTasks.objects.filter(projects=project).values(
        'projects', 'scope', 'category', 'title', 'rev', 'task_benchmark'
    )

    # For unique approved hours
    unique_keys = set()
    approved_hours = 0
    for task in tasks:
        key = (task['projects'], task['scope'], task['category'], task['title'], task['rev'])
        if key not in unique_keys:
            unique_keys.add(key)
            approved_hours += float(task['task_benchmark'] or 0)

    # Total worktime for this project
    total_worktime = TrackerTasks.objects.filter(projects=project).aggregate(
        total=Sum('time')
    )['total'] or 0

    return JsonResponse({
        "projects": [project],
        "approvedHours": [approved_hours],
        "totalWorktime": [total_worktime]
    })


from django.http import JsonResponse
from .models import TrackerTasks

def get_task_data(request, project):
    tasks = TrackerTasks.objects.filter(projects=project)

    task_names = []
    approved_hours_list = []
    total_worktime_list = []
    user_worktime_map = {}

    for task in tasks:
        task_name = task.title
        task_names.append(task_name)
        approved_hours_list.append(float(task.task_benchmark or 0))
        total_worktime_list.append(float(task.time or 0))

        user = task.assigned or "Unassigned"
        user_worktime_map[user] = user_worktime_map.get(user, 0) + float(task.time or 0)

    return JsonResponse({
        "tasks": task_names,
        "approvedHours": approved_hours_list,
        "totalWorktime": total_worktime_list,
        "userWorktimes": user_worktime_map
    })



# from django.http import JsonResponse
# from .models import TrackerTasks
# from django.db.models import Sum

# def get_team_chart_datas(request):
#     # Fetch all tasks from the database
#     team_data = TrackerTasks.objects.values(
#         'team',
#         'projects',
#         'scope',
#         'category',
#         'title',
#         'rev',
#         'd_no',  # Include d_no in the values
#         'task_benchmark'
#     ).order_by('team')

#     # Dictionary to accumulate approved hours and worktime
#     approved_hours_dict = {}
#     total_worktime_dict = {}

#     # Loop through the data to manually filter duplicates for Approved Hours
#     unique_combinations = set()
#     for entry in team_data:
#         # Create a unique key based on the combination of team, project, scope, category, title, rev, and d_no
#         key = (entry['team'], entry['projects'], entry['scope'], 
#                entry['category'], entry['title'], entry['rev'], entry['d_no'])
        
#         # Check if the combination already exists
#         if key not in unique_combinations:
#             unique_combinations.add(key)  # Mark this combination as counted
            
#             # Initialize the dictionary if team is not present
#             if entry['team'] not in approved_hours_dict:
#                 approved_hours_dict[entry['team']] = 0
            
#             # Add the benchmark value (approved hours) for the unique combination
#             approved_hours_dict[entry['team']] += float(entry['task_benchmark'] or 0)
    
#     # Calculate Total Worktime by filtering just the Team
#     worktime_data = TrackerTasks.objects.values('team').annotate(
#         total_worktime=Sum('time')
#     ).order_by('team')

#     # Populate the dictionary with worktime values
#     for entry in worktime_data:
#         total_worktime_dict[entry['team']] = entry['total_worktime']

#     # Prepare data for JSON response
#     data = {
#         "teams": list(approved_hours_dict.keys()),
#         "approvedHours": list(approved_hours_dict.values()),
#         "totalWorktime": [total_worktime_dict.get(team, 0) for team in approved_hours_dict.keys()]
#     }

#     return JsonResponse(data)


def get_project(request, team):
    projects = TrackerTasks.objects.filter(team=team).values('projects').distinct()
    project_names = [project['projects'] for project in projects]
    return JsonResponse({"projects": project_names})


from django.http import JsonResponse
from .models import TrackerTasks
from django.db.models import Sum

def get_task_datas(request, project):  # ✅ Only project now
    # Fetch all unique task combinations under the project
    tasks = TrackerTasks.objects.filter(projects=project).values(
        'projects', 'scope', 'category', 'title', 'rev', 'd_no'
    ).distinct()

    approved_hours = []
    total_worktime = []
    task_titles = []
    user_worktimes = {}

    for task in tasks:
        task_titles.append(task['title'])

        filters = {
            'projects': project,
            'title': task['title'],
            'rev': task['rev'],
            'd_no': task['d_no'],
        }

        # Approved hours
        approved_sum = TrackerTasks.objects.filter(**filters).aggregate(
            total=Sum('task_benchmark')
        )['total'] or 0
        approved_hours.append(approved_sum)

        # Total worktime
        worktime_sum = TrackerTasks.objects.filter(**filters).aggregate(
            total=Sum('time')
        )['total'] or 0
        total_worktime.append(worktime_sum)

    # Aggregate user worktimes under this project
    user_data = TrackerTasks.objects.filter(projects=project).values('assigned').annotate(
        total_worktime=Sum('time')
    )

    for entry in user_data:
        username = entry['assigned'] or 'Unassigned'
        if entry['total_worktime'] is not None:
            user_worktimes[username] = entry['total_worktime']

    return JsonResponse({
        "tasks": task_titles,
        "approvedHours": approved_hours,
        "totalWorktime": total_worktime,
        "userWorktimes": user_worktimes
    })


from django.http import JsonResponse
from .models import TrackerTasks

def get_project_data(request, project):
    # Fetch all tasks for the specified project
    project_data = TrackerTasks.objects.filter(projects=project).values(
        'projects',
        'scope',
        'category',
        'title',
        'rev',
        'd_no',
        'task_benchmark',
        'time'
    )

    # Use dicts to accumulate approved hours and total worktime
    approved_hours = 0
    total_worktime = 0
    unique_combinations = set()

    for entry in project_data:
        key = (
            entry['projects'],
            entry['scope'],
            entry['category'],
            entry['title'],
            entry['rev'],
            entry['d_no']
        )

        # Add approved hours only once for each unique key
        if key not in unique_combinations:
            unique_combinations.add(key)
            approved_hours += float(entry['task_benchmark'] or 0)

        # Always sum total worktime
        total_worktime += float(entry['time'] or 0)

    # Prepare JSON response
    data = {
        "projects": [project],
        "approvedHours": [approved_hours],
        "totalWorktime": [total_worktime]
    }

    return JsonResponse(data)


from django.shortcuts import render
from django.http import JsonResponse
from .models import TrackerTasks

# Assuming user_data is a global variable

def get_projects_data(request):
    # Fetch distinct project names from the TrackerTasks model
    projects = TrackerTasks.objects.values('projects').distinct()

    # Convert queryset to list of project names
    projects = [project['projects'] for project in projects]

    # Return the data as JSON
    return JsonResponse({
        'projects': projects
    })

from django.shortcuts import render
from django.http import JsonResponse
from .models import TrackerTasks
from .forms import ProjectStatusUpdateForm

# Assuming user_data is a global variable

def update_project_status(request):
    if request.method == 'POST':
        form = ProjectStatusUpdateForm(request.POST)
        if form.is_valid():
            # Get the selected data from the form
            project_name = form.cleaned_data['projects']
            project_status = form.cleaned_data['project_status']

            # Update only the selected project's status
            tasks = TrackerTasks.objects.filter(projects=project_name)

            # If tasks exist, update their project_status
            if tasks.exists():
                tasks.update(project_status=project_status)
                return JsonResponse({"success": "Project status updated successfully"}, status=200)
            else:
                return JsonResponse({"error": "Project not found"}, status=404)

    else:
        # Display the form initially (GET request)
        form = ProjectStatusUpdateForm()

    return render(request, 'project_tracker.html', {'form': form})


from django.shortcuts import render
from django.http import JsonResponse
from .models import TeamRanking
import json
from django.views.decorators.csrf import csrf_exempt

def team_ranking_page(request):
    return render(request, 'project_tracker.html')
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import TeamRanking
import json

@csrf_exempt
def add_team_ranking(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        
        # Check if the team member already exists in the specified team
        existing_record = TeamRanking.objects.filter(
            team_name=data['team_name'],
            team_member=data['team_member']
        ).first()
        
        if existing_record:
            # Update the existing record
            existing_record.speed_of_execution = data['speed_of_execution']
            existing_record.complaints_of_check_list = data['complaints_of_check_list']
            existing_record.task_ownership = data['task_ownership']
            existing_record.understanding_task = data['understanding_task']
            existing_record.quality_of_work = data['quality_of_work']
            existing_record.save()
            return JsonResponse({'status': 'updated'})
        else:
            # Create a new record if it doesn't exist
            TeamRanking.objects.create(
                team_name=data['team_name'],
                team_member=data['team_member'],
                speed_of_execution=data['speed_of_execution'],
                complaints_of_check_list=data['complaints_of_check_list'],
                task_ownership=data['task_ownership'],
                understanding_task=data['understanding_task'],
                quality_of_work=data['quality_of_work'],
            )
            return JsonResponse({'status': 'created'})
    return JsonResponse({'status': 'invalid request'}, status=400)


# In your views.py
@csrf_exempt
def get_team_member_details(request):
    team_name = request.GET.get('team_name')
    team_member = request.GET.get('team_member')
    
    data = TeamRanking.objects.filter(team_name=team_name, team_member=team_member).values().first()
    return JsonResponse(data, safe=False)


from django.http import JsonResponse
from .models import TeamRanking

def get_team_rankings(request):
    # Fetch the team data ordered by date in descending order (newest first)
    team_data = list(TeamRanking.objects.order_by('-date').values())
    return JsonResponse(team_data, safe=False)



from django.http import JsonResponse
from .models import TrackerTasks

def get_team_names(request):
    team_data = (
        TrackerTasks.objects.values('team', 'assigned')
        .distinct()
        .exclude(team__isnull=True, assigned__isnull=True)
    )
    formatted_data = [{'team_name': item['team'], 'team_member': item['assigned']} for item in team_data]
    return JsonResponse(formatted_data, safe=False)



from django.http import JsonResponse
from django.core.mail import send_mail
from django.conf import settings
import json

def send_notification(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            message = data.get('message', '')
            recipient = data.get('recipient', '')
            print(recipient)  
            if not message or not recipient:
                return JsonResponse({"error": "Invalid data"}, status=400)

            send_mail(
                subject="Task Benchmark Missing",  # Email subject
                message=message,                   # The message body
                from_email=settings.DEFAULT_FROM_EMAIL,  # From address
                recipient_list=[recipient],
                       # Recipient's email
                fail_silently=False
            )
            return JsonResponse({"success": "Notification sent to admin."})
        except Exception as e:
            print(f"Failed to send email: {e}")
            return JsonResponse({"error": f"Failed to send notification: {str(e)}"}, status=500)
    else:
        return JsonResponse({"error": "Invalid request method"}, status=405)

